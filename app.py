"""
GTL Control Room - Flask backend.

Job 1: login/logout, session-based.
Job 2: serve the dashboard shell (any logged-in role).
Job 3: /api/dashboard-data - the ONLY place role filtering matters.
       Every byte sent to the browser passes through
       permissions.filter_payload_for_role() first. A client-role
       session literally cannot receive tampering data in the HTTP
       response, not "won't show it", cannot.
Job 4: /api/export/<name> - gated by permissions.EXPORT_ACCESS.

Run locally:
    python app.py
Then open http://127.0.0.1:5000
"""

import os
import re
import sys
import json
import time
import hmac
import base64
import threading
import functools
from datetime import datetime
from flask import Flask, request, session, jsonify, redirect, url_for, Response, render_template

from dotenv import load_dotenv
load_dotenv()  # loads .env for local dev; no-op on Render, which injects real env vars directly

import users as users_store
from users import verify_login, DEACTIVATED
import permissions
from permissions import (
    filter_payload_for_role, allowed_panels, EXPORT_ACCESS,
    MANAGE_USERS_ROLES, ROLES, PANEL_LABELS, LockoutError,
)
from respond_tokens import make_respond_token, read_respond_token

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "importer"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "fleet_logic"))

from schema import now_eat
import atomic_json

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-only-change-me")


def public_base_url():
    """
    The base URL for links that leave the app in an email, which must
    never depend on which machine happened to be running Flask when the
    notification was sent. request.url_root reflects THAT REQUEST's
    host - correct for a real visitor Browse the deployed site, but
    "http://localhost:5000/" for any local dev run or test script, and a
    link like that is dead for anyone who isn't the exact machine that
    sent it. PUBLIC_BASE_URL is an explicit override; RENDER_EXTERNAL_URL
    is injected automatically by Render for every web service, so
    production needs no manual configuration at all. request.url_root is
    the last resort, kept only so interactive local testing still works.

    A real production incident: PUBLIC_BASE_URL got set to
    "http://localhost:5000" on the actual live Render service (almost
    certainly copied in from a local .env at some point) - since it's
    checked FIRST, that silently overrode the otherwise-correct
    request.url_root fallback for every real visitor, and every email
    link went out dead. _is_probably_local_leftover() below refuses to
    trust an override that looks like a dev-machine artifact rather
    than a real public domain, so a misconfigured env var can't take
    priority over the live request it's actually being sent from.
    """
    configured = os.environ.get("PUBLIC_BASE_URL") or os.environ.get("RENDER_EXTERNAL_URL")
    if configured and not _is_probably_local_leftover(configured):
        return configured.rstrip("/")
    if configured:
        print(f"WARNING: ignoring PUBLIC_BASE_URL/RENDER_EXTERNAL_URL={configured!r} - "
              f"looks like a local-dev leftover, not a real public domain. Falling back "
              f"to the current request's own host instead.")
    return request.url_root.rstrip("/")


def _is_probably_local_leftover(url):
    """True for exactly the kind of value that should never win over a
    real incoming request's own host: localhost, 127.0.0.1, 0.0.0.0, or
    port 5000/8000-style dev-server ports."""
    lowered = url.lower()
    return any(marker in lowered for marker in ("localhost", "127.0.0.1", "0.0.0.0"))


def _notify_async(plate, comment, added_by, role, entry_type, requires_followup, respond_urls):
    """
    Fires notifications.on_comment_added() on a background thread. The
    Sheets write it follows is already durable by the time this runs -
    email is a courtesy the requester was never waiting on, so a slow or
    blocked SMTP host (see mailer.py's port 587->465 fallback, which by
    itself can take several seconds) must not hold the HTTP response
    open. Errors are only logged here; there is no request left to
    report them to.
    """
    def _run():
        try:
            import notifications
            result = notifications.on_comment_added(
                plate, comment, added_by, role, entry_type, requires_followup, respond_urls)
            if not result["sent"]:
                print(f"Notification email not sent for {plate}: {result['reason']}")
        except Exception as e:
            print(f"Notification email failed for {plate}: {e}")
    threading.Thread(target=_run, daemon=True).start()


DATA_PATH = os.path.join(os.path.dirname(__file__), "data", "fleet_today.json")


def load_dashboard_data():
    if not os.path.exists(DATA_PATH):
        return {}
    with open(DATA_PATH) as f:
        return json.load(f)


def login_required(fn):
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return jsonify({"error": "not authenticated"}), 401
        return fn(*args, **kwargs)
    return wrapper


@app.route("/api/login", methods=["POST"])
def api_login():
    body = request.get_json(force=True, silent=True) or {}
    username = body.get("username", "")
    password = body.get("password", "")
    role = verify_login(username, password)
    if role is None:
        return jsonify({"error": "Invalid username or password"}), 401
    if role == DEACTIVATED:
        return jsonify({"error": "This account has been deactivated. Contact your administrator."}), 403
    session["username"] = username
    session["role"] = role
    return jsonify({"username": username, "role": role, "panels": allowed_panels(role)})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/session")
def api_session():
    if "username" not in session:
        return jsonify({"authenticated": False})
    return jsonify({
        "authenticated": True,
        "username": session["username"],
        "role": session["role"],
        "panels": allowed_panels(session["role"]),
    })


@app.route("/api/users", methods=["GET"])
@login_required
def api_list_users():
    """
    Admin and technician accounts can see who has a login and what role
    they hold - never password hashes, those never leave users.json.
    Client role is deliberately excluded: it's an external account
    (GTL themselves), not part of your team's account administration.
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    all_users = users_store.load_users()
    return jsonify({
        "users": [
            {
                "username": name,
                "role": info.get("role", ""),
                # The address notifications go to. Blank means this person
                # simply isn't reachable by email yet - the UI says so rather
                # than letting someone assume they're being kept informed.
                "email": info.get("email", ""),
                # A user can be assigned more than one client, so this is
                # always a list even when there's only one today.
                "clients": info.get("clients", []),
                # Blank for anyone who hasn't signed in since last-login
                # tracking was added - the UI shows "Never" rather than
                # inventing a date that was never recorded.
                "lastLogin": info.get("last_login", ""),
                "createdAt": info.get("created_at", ""),
                "active": info.get("active", True),
            }
            for name, info in sorted(all_users.items())
        ],
        # Tells the UI whether accounts are in the durable store or the
        # ephemeral local fallback, so it can warn rather than quietly
        # letting someone create accounts that a deploy will erase.
        "durable": users_store._sheets_available(),
    })


@app.route("/api/users", methods=["POST"])
@login_required
def api_create_user():
    """
    Lets admin and technician accounts provision new logins from the
    UI instead of needing shell access to run users.py by hand - the
    same add_user() call, just reachable without a terminal. Refuses
    to silently overwrite an existing username: that's still a job for
    users.py directly, a deliberate, explicit action, not a UI click.
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    body = request.get_json(force=True, silent=True) or {}
    username = (body.get("username") or "").strip()
    role = (body.get("role") or "").strip()
    password = body.get("password") or ""
    email = (body.get("email") or "").strip()
    clients_raw = body.get("clients") or []
    if isinstance(clients_raw, str):
        clients_raw = [c.strip() for c in clients_raw.split(",")]
    clients = [str(c).strip() for c in clients_raw if str(c).strip()]

    if not username or not role or not password:
        return jsonify({"error": "'username', 'role', and 'password' are all required"}), 400
    if role not in ROLES:
        return jsonify({"error": f"'role' must be one of {list(ROLES)}"}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters"}), 400
    if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify({"error": "That doesn't look like a valid email address"}), 400
    if username in users_store.load_users():
        return jsonify({"error": f"'{username}' already has an account"}), 409

    # A role that sees every client already gets everything - selecting
    # clients for it would be meaningless and is how an admin account
    # ended up scoped to three clients by mistake. Force it clear rather
    # than trusting whatever the form sent.
    if permissions.sees_all_clients(role):
        clients = []
    elif not clients:
        return jsonify({"error": "Select at least one client for this account - "
                                 "without one, this role can't see any fleet data at all."}), 400

    # Nobody can hand out access they don't hold themselves. A
    # technician scoped to GTL must not be able to create an account
    # with access to AGL - that would be a trivial privilege escalation
    # (create the account, then log in as it). Only admin, who holds
    # every client, passes this unconditionally.
    ok, disallowed = permissions.can_grant_clients(
        session["role"], _assigned_clients(session["username"]), clients)
    if not ok:
        return jsonify({"error": "You can only grant access to clients you have access to yourself. "
                                 f"Not yours to grant: {', '.join(disallowed)}"}), 403
    # A role that sees every client is likewise not something a
    # non-admin can create, for the same escalation reason.
    if permissions.sees_all_clients(role) and not permissions.sees_all_clients(session["role"]):
        return jsonify({"error": f"Only an admin can create a '{role}' account."}), 403

    try:
        users_store.add_user(username, role, password, clients, email)
    except Exception as e:
        return jsonify({"error": f"Could not save the account: {e}"}), 503
    return jsonify({"username": username, "role": role, "clients": clients, "email": email}), 201


@app.route("/api/users/<username>/email", methods=["PUT"])
@login_required
def api_set_user_email(username):
    """Backfills an address for an account created before emails were
    collected, so existing users become reachable without being recreated."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    body = request.get_json(force=True, silent=True) or {}
    email = (body.get("email") or "").strip()
    if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify({"error": "That doesn't look like a valid email address"}), 400

    try:
        found = users_store.set_email(username, email)
    except Exception as e:
        return jsonify({"error": f"Could not save the address: {e}"}), 503
    if not found:
        return jsonify({"error": f"No account called '{username}'"}), 404
    return jsonify({"username": username, "email": email})


@app.route("/api/users/<username>/role", methods=["PUT"])
@login_required
def api_set_user_role(username):
    """Changes a user's role. Same escalation rule as creating an
    account: nobody can hand out a role with more reach than their own."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    target = users_store.get_user(username)
    if not target:
        return jsonify({"error": f"No account called '{username}'"}), 404

    body = request.get_json(force=True, silent=True) or {}
    role = (body.get("role") or "").strip()
    if role not in ROLES:
        return jsonify({"error": f"'role' must be one of {list(ROLES)}"}), 400
    if permissions.sees_all_clients(role) and not permissions.sees_all_clients(session["role"]):
        return jsonify({"error": f"Only an admin can grant the '{role}' role."}), 403
    if username == session["username"] and not permissions.sees_all_clients(role):
        return jsonify({"error": "You can't demote your own account while logged in as it."}), 400

    try:
        users_store.set_role(username, role)
        # A role that sees every client needs no client list of its own -
        # clear it rather than leaving a stale, meaningless scope behind
        # (this is exactly how an admin account ended up tagged with
        # three clients it never needed and couldn't be filtered by).
        if permissions.sees_all_clients(role):
            users_store.set_clients(username, [])
    except Exception as e:
        return jsonify({"error": f"Could not save the role: {e}"}), 503
    return jsonify({"username": username, "role": role})


@app.route("/api/users/<username>/active", methods=["PUT"])
@login_required
def api_set_user_active(username):
    """Suspends or restores a login without deleting the account."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    if username == session["username"]:
        return jsonify({"error": "You can't deactivate your own account while logged in as it."}), 400

    body = request.get_json(force=True, silent=True) or {}
    active = bool(body.get("active", True))

    try:
        found = users_store.set_active(username, active)
    except Exception as e:
        return jsonify({"error": f"Could not save that change: {e}"}), 503
    if not found:
        return jsonify({"error": f"No account called '{username}'"}), 404
    return jsonify({"username": username, "active": active})


@app.route("/api/users/<username>/clients", methods=["PUT"])
@login_required
def api_set_user_clients(username):
    """Updates the client assignments for a user."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    target = users_store.get_user(username)
    if not target:
        return jsonify({"error": f"No account called '{username}'"}), 404

    body = request.get_json(force=True, silent=True) or {}
    clients_raw = body.get("clients") or []
    if isinstance(clients_raw, str):
        clients_raw = [c.strip() for c in clients_raw.split(",")]
    clients = [str(c).strip() for c in clients_raw if str(c).strip()]

    # A role that already sees every client (admin) has no use for a
    # client list - force it clear instead of storing (or rejecting) a
    # selection that would just be dead weight.
    if permissions.sees_all_clients(target.get("role", "")):
        clients = []
    elif not clients:
        return jsonify({"error": "Select at least one client for this account - "
                                 "without one, this role can't see any fleet data at all."}), 400
    else:
        # Users cannot grant clients they don't have access to themselves
        ok, disallowed = permissions.can_grant_clients(
            session["role"], _assigned_clients(session["username"]), clients)
        if not ok:
            return jsonify({"error": "You can only assign clients you have access to yourself. "
                                     f"Not yours to assign: {', '.join(disallowed)}"}), 403

    try:
        found = users_store.set_clients(username, clients)
    except Exception as e:
        return jsonify({"error": f"Could not save the clients: {e}"}), 503
    if not found:
        return jsonify({"error": f"No account called '{username}'"}), 404
    return jsonify({"username": username, "clients": clients})


@app.route("/api/users/<username>", methods=["DELETE"])
@login_required
def api_delete_user(username):
    """
    There was previously no way to remove an account short of editing
    the Sheet by hand. Blocks deleting your own account specifically -
    every other account is fair game, but locking yourself out by
    accident (the one mistake with no recovery path from inside the
    app) is worth a hard stop rather than a confirm dialog alone.
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    if username == session["username"]:
        return jsonify({"error": "You can't delete your own account while logged in as it."}), 400

    try:
        found = users_store.delete_user(username)
    except Exception as e:
        return jsonify({"error": f"Could not delete the account: {e}"}), 503
    if not found:
        return jsonify({"error": f"No account called '{username}'"}), 404
    return jsonify({"ok": True, "username": username})


def _valid_emails(raw):
    """raw is a comma-separated string from the UI. Returns
    (clean_list, error_message) - error_message is None when every
    address present is well-formed, so a single typo doesn't silently
    drop the rest of a client's contacts."""
    emails = [e.strip() for e in raw.split(",") if e.strip()]
    bad = [e for e in emails if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", e)]
    if bad:
        return None, f"These don't look like valid email addresses: {', '.join(bad)}"
    return emails, None


@app.route("/api/clients", methods=["GET"])
@login_required
def api_list_clients():
    """
    The clients this session may see - not all of them.

    This previously returned every client to any logged-in role,
    including their contact email addresses and (since the mapping
    build) their platform account ids. Those are another company's
    business contacts, so a client-role login for one organisation
    could read the operational contacts of every other organisation on
    the platform.

    Admin still sees everything, which is what makes the mapping UI
    work; everyone else sees only their own.
    """
    try:
        import sheets_store
        clients = sheets_store.load_clients()
    except Exception as e:
        return jsonify({"error": f"Could not load clients: {e}"}), 503
    allowed = _visible_clients_for_session()
    if allowed is not None:
        clients = [c for c in clients if c.get("name", "").strip() in allowed]
    return jsonify({"clients": clients})


@app.route("/api/platform-accounts", methods=["GET"])
@login_required
def api_platform_accounts():
    """
    Every account visible on each platform, normalised to {id, name},
    for the admin UI's mapping dropdowns. Picking a name is the whole
    point: the raw identifiers are a 19-digit MiX group id, a
    base64-looking Teletrac ClientId and a 19-digit FT fleetId, none of
    which can be eyeballed for correctness, and mapping a client to the
    wrong one silently shows them another company's vehicles.

    Each platform is fetched independently and reports its own error, so
    one unreachable API leaves the other two dropdowns usable instead of
    failing the whole dialog.
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    out = {"mix": [], "teletrac": [], "ftCloud": [], "errors": {}}

    try:
        from adapters.mix_api_client import MixApiClient
        client = MixApiClient()
        if not client.is_configured():
            raise RuntimeError("MiX API credentials not set")
        out["mix"] = sorted(
            ({"id": str(g.get("GroupId")), "name": g.get("Name") or str(g.get("GroupId"))}
             for g in client.get_organisation_groups()),
            key=lambda g: g["name"].lower())
    except Exception as e:
        out["errors"]["mix"] = str(e)

    try:
        from adapters.teletrac_api_client import TeletracApiClient
        client = TeletracApiClient()
        if not client.is_configured():
            raise RuntimeError("Teletrac API credentials not set")
        out["teletrac"] = sorted(
            ({"id": str(c.get("ClientId")), "name": c.get("vCompanyName") or str(c.get("ClientId"))}
             for c in client.get_all_clients()),
            key=lambda c: c["name"].lower())
    except Exception as e:
        out["errors"]["teletrac"] = str(e)

    try:
        from adapters.ft_cloud_api_client import FtCloudApiClient
        client = FtCloudApiClient()
        if not client.is_configured():
            raise RuntimeError("FT Cloud API credentials not set")
        out["ftCloud"] = sorted(
            ({"id": str(f.get("fleetId")), "name": f.get("fleetName") or str(f.get("fleetId"))}
             for f in client.get_fleets()),
            key=lambda f: f["name"].lower())
    except Exception as e:
        out["errors"]["ftCloud"] = str(e)

    return jsonify(out)


def _platform_ids(body):
    """Pulls the three platform-id lists out of a request body,
    accepting either a list or a single string for each."""
    def as_list(value):
        if value is None:
            return None          # not supplied - leave unchanged on update
        if isinstance(value, str):
            value = [value]
        return [str(v).strip() for v in value if str(v).strip()]
    return (as_list(body.get("mixOrgIds")),
            as_list(body.get("teletracClientIds")),
            as_list(body.get("ftCloudFleetIds")))


@app.route("/api/clients", methods=["POST"])
@login_required
def api_create_client():
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    body = request.get_json(force=True, silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "'name' is required"}), 400
    emails, err = _valid_emails(body.get("emails") or "")
    if err:
        return jsonify({"error": err}), 400
    mix, teletrac, ft = _platform_ids(body)

    # A platform account belongs to exactly one client. Letting two
    # clients claim the same MiX org would put the same vehicles under
    # both names and make the per-client access boundary meaningless -
    # and classify_fleet would log a conflict for every affected plate.
    conflict = _platform_account_conflict(name, mix, teletrac, ft)
    if conflict:
        return jsonify({"error": conflict}), 409

    try:
        import sheets_store
        sheets_store.add_client(name, emails, mix or [], teletrac or [], ft or [])
    except ValueError as e:
        return jsonify({"error": str(e)}), 409
    except Exception as e:
        return jsonify({"error": f"Could not save the client: {e}"}), 503
    _invalidate_client_registry()
    return jsonify({"name": name, "emails": emails, "mixOrgIds": mix or [],
                    "teletracClientIds": teletrac or [], "ftCloudFleetIds": ft or []}), 201


def _platform_account_conflict(name, mix, teletrac, ft):
    """Returns a human-readable message if any of these platform
    accounts is already mapped to a DIFFERENT client, else None."""
    try:
        import sheets_store
        existing = sheets_store.load_clients()
    except Exception:
        return None  # can't check - don't block the write on it
    for other in existing:
        if other["name"].strip().lower() == name.strip().lower():
            continue
        for requested, key, label in ((mix, "mixOrgIds", "MiX organisation"),
                                      (teletrac, "teletracClientIds", "Teletrac client"),
                                      (ft, "ftCloudFleetIds", "FT Cloud fleet")):
            clash = set(requested or []) & set(other.get(key) or [])
            if clash:
                return (f"That {label} is already mapped to '{other['name']}'. "
                        f"A platform account can only belong to one client.")
    return None


def _invalidate_client_registry():
    """The pollers memoise the registry for a few minutes; after an edit
    they should pick it up on the next cycle, not minutes later."""
    try:
        import client_registry
        client_registry.invalidate_cache()
    except Exception as e:
        print(f"Could not invalidate the client registry cache: {e}")


@app.route("/api/clients/<name>/platforms", methods=["PUT"])
@login_required
def api_set_client_platforms(name):
    """
    Remaps which platform accounts belong to this client. Omitting a
    platform leaves it unchanged; sending an empty list clears it,
    which is a different and meaningful instruction ("this client is no
    longer on that platform").
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    body = request.get_json(force=True, silent=True) or {}
    mix, teletrac, ft = _platform_ids(body)
    if mix is None and teletrac is None and ft is None:
        return jsonify({"error": "Nothing to update - supply at least one platform's ids"}), 400

    conflict = _platform_account_conflict(name, mix, teletrac, ft)
    if conflict:
        return jsonify({"error": conflict}), 409

    try:
        import sheets_store
        found = sheets_store.set_client_platforms(name, mix, teletrac, ft)
    except Exception as e:
        return jsonify({"error": f"Could not save: {e}"}), 503
    if not found:
        return jsonify({"error": f"No client called '{name}'"}), 404
    _invalidate_client_registry()
    return jsonify({"name": name, "mixOrgIds": mix, "teletracClientIds": teletrac,
                    "ftCloudFleetIds": ft})


@app.route("/api/clients/<name>/emails", methods=["PUT"])
@login_required
def api_set_client_emails(name):
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    body = request.get_json(force=True, silent=True) or {}
    emails, err = _valid_emails(body.get("emails") or "")
    if err:
        return jsonify({"error": err}), 400
    try:
        import sheets_store
        found = sheets_store.set_client_emails(name, emails)
    except Exception as e:
        return jsonify({"error": f"Could not save: {e}"}), 503
    if not found:
        return jsonify({"error": f"No client called '{name}'"}), 404
    return jsonify({"name": name, "emails": emails})


@app.route("/api/clients/<name>", methods=["DELETE"])
@login_required
def api_delete_client(name):
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    try:
        import sheets_store
        found = sheets_store.delete_client(name)
    except Exception as e:
        return jsonify({"error": f"Could not delete: {e}"}), 503
    if not found:
        return jsonify({"error": f"No client called '{name}'"}), 404
    # Its platform accounts are no longer claimed by anyone - stop the
    # pollers fetching them on the next cycle rather than minutes later.
    _invalidate_client_registry()
    return jsonify({"ok": True})


@app.route("/api/role-panels", methods=["GET"])
@login_required
def api_role_panels():
    """The current panel->roles matrix. Every cell is freely tickable -
    the only rule the server still enforces is that Manage Users can't
    be taken away from every role at once, which is checked on save."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    return jsonify({
        "roles": list(ROLES),
        "panels": [
            {
                "id": panel,
                "label": PANEL_LABELS.get(panel, panel),
                "roles": list(roles),
            }
            for panel, roles in permissions.PANEL_ACCESS.items()
        ],
    })


@app.route("/api/role-panels", methods=["POST"])
@login_required
def api_save_role_panels():
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    body = request.get_json(force=True, silent=True) or {}
    new_access = body.get("panels")
    if not isinstance(new_access, dict):
        return jsonify({"error": "'panels' must be an object of panelId -> [roles]"}), 400

    try:
        updated = permissions.save_panel_access(new_access)
    except LockoutError as e:
        return jsonify({"error": str(e)}), 400
    except OSError as e:
        return jsonify({"error": f"Could not save role settings: {e}"}), 503

    return jsonify({"ok": True, "panels": {p: list(r) for p, r in updated.items()}})


def _overlay_feedback(raw):
    """
    Applies the human layer on top of the imported telemetry. Returns
    (payload, overlay_ok) - on failure the telemetry still goes out and
    the caller flags the payload as stale, because a dashboard that
    silently drops feedback is worse than one that admits it couldn't
    load it.
    """
    try:
        import sheets_store
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), "fleet_logic"))
        from feedback_overlay import apply_feedback
        return apply_feedback(raw, sheets_store.load_feedback_cached()), True
    except Exception as e:
        print(f"Feedback overlay unavailable this request: {e}")
        return raw, False


_user_clients_cache = {}          # username -> (fetched_at, clients list)
_USER_CLIENTS_TTL_SECONDS = 30


def _assigned_clients(username):
    """
    This account's client assignments, cached briefly.

    users.load_users() is an uncached full Sheets read, and the
    dashboard now re-fetches itself every 60s in every open tab, so
    reading it per request would put a Sheets round trip on the hot
    path. 30 seconds keeps that cheap while still making an access
    change (or revocation) take effect almost immediately.

    On a lookup failure the last known good value is reused if there is
    one, and otherwise access is denied rather than granted - a Sheets
    outage must not turn into a moment where a restricted account can
    see every client's data.
    """
    now = time.time()
    cached = _user_clients_cache.get(username)
    if cached and now - cached[0] < _USER_CLIENTS_TTL_SECONDS:
        return cached[1]
    try:
        user = users_store.get_user(username) or {}
        clients = user.get("clients", [])
        _user_clients_cache[username] = (now, clients)
        return clients
    except Exception as e:
        if cached:
            print(f"Could not refresh client access for {username} ({e}); using the last known assignment")
            return cached[1]
        print(f"Could not determine client access for {username} ({e}); denying access this request")
        return []


_plate_client_cache = {"mtime": None, "map": {}}


def _plate_client_map():
    """
    {plate: client} for every asset in the current dashboard payload.

    Needed wherever a record identifies a vehicle by plate but carries
    no client of its own - feedback comments, the tampering sections -
    since those still have to be scoped to what the session may see.
    Keyed off fleet_today.json's mtime so it rebuilds when the pollers
    rewrite it and costs nothing on every other request.
    """
    try:
        mtime = os.path.getmtime(DATA_PATH)
    except OSError:
        return {}
    if _plate_client_cache["mtime"] == mtime:
        return _plate_client_cache["map"]
    mapping = {
        str(r.get("plate", "")).strip(): str(r.get("client", "")).strip()
        for r in (load_dashboard_data().get("full") or [])
        if isinstance(r, dict) and r.get("plate")
    }
    _plate_client_cache.update({"mtime": mtime, "map": mapping})
    return mapping


def _scoped_snapshot(path, group_key):
    """
    A poll snapshot with other clients' accounts removed.

    These files are the raw platform payload - every plate, position and
    device id for every organisation polled - so serving them whole
    handed a scoped technician the complete fleet of every client, in a
    form the dashboard's own filtering never touched. Each group's rows
    carry the client that owns them (see the pollers), so the groups are
    filtered by that rather than by trying to re-derive ownership here.
    """
    if not os.path.exists(path):
        return jsonify({"status": "no_snapshot_yet"})
    try:
        with open(path) as f:
            snapshot = json.load(f)
    except (ValueError, OSError) as e:
        return jsonify({"error": f"Snapshot unreadable: {e}"}), 503

    allowed = _visible_clients_for_session()
    if allowed is None:
        return jsonify(snapshot)

    groups = snapshot.get(group_key) or {}
    kept = {
        gid: rows for gid, rows in groups.items()
        if any(str(r.get("client", "")).strip() in allowed for r in rows if isinstance(r, dict))
    }
    scoped = dict(snapshot)
    scoped[group_key] = kept
    scoped["counts"] = {gid: len(rows) for gid, rows in kept.items()}
    # The id lists and the id->client map name every organisation
    # polled, so they leak the client roster even with the rows gone.
    for key in ("orgIds", "clientIds", "fleetIds"):
        if key in scoped:
            scoped[key] = [i for i in scoped[key] if str(i) in kept]
    for key in ("clientByOrgId", "clientByClientId", "clientByFleetId"):
        if key in scoped:
            scoped[key] = {k: v for k, v in scoped[key].items() if v in allowed}
    # Retired plates are fleet-wide and not attributable per client here.
    scoped.pop("decommissionedPlates", None)
    return jsonify(scoped)


def _plate_allowed(plate):
    """
    Whether this session may see or act on one vehicle.

    Every endpoint that takes a plate from the request needs this:
    without it the plate is a free-form parameter that reaches straight
    into another client's data, and the dashboard's filtering counts for
    nothing because the underlying record is still one URL away.

    An unknown plate is refused rather than allowed. It can't be
    attributed to a client, so permitting it would leave a hole for
    exactly the plates whose ownership is unclear.
    """
    allowed = _visible_clients_for_session()
    if allowed is None:
        return True
    return _plate_client_map().get(str(plate).strip(), "") in allowed


def _deny_plate(plate):
    """Deliberately the same response for 'not yours' and 'no such
    vehicle': distinguishing them would let anyone enumerate which
    plates exist on the platform by watching the status code."""
    return jsonify({"error": f"No vehicle '{plate}' available to your account"}), 404


def _visible_clients_for_session():
    """None = unrestricted (admin). A set = exactly those clients."""
    role = session["role"]
    if permissions.sees_all_clients(role):
        return None
    return permissions.visible_clients(role, _assigned_clients(session["username"]))


def _client_filter_options(payload, allowed):
    """
    What the header's client filter offers.

    Sourced from the REGISTRY, not from the rows in this payload. Those
    are different sets and the difference matters: a client that has
    just been added, whose vehicles haven't been polled yet, or whose
    platform fetch is failing, has no rows - and deriving the list from
    rows made it silently absent from the filter with nothing on screen
    explaining why. Listing it and showing an honest empty dashboard is
    the better failure.

    Union'd with the clients actually present, so a client with rows but
    no registry entry (a legacy row, or an unmapped platform account
    landing under "Unassigned") still shows up rather than becoming
    unreachable data.

    This exposes client NAMES only, and only ones the session is
    entitled to - `allowed` is applied to both halves. It is not a way
    into another client's data; the row filtering above is.
    """
    present = {
        str(r.get("client", "")).strip()
        for section in permissions.CLIENT_SCOPED_SECTIONS
        for r in (payload.get(section) or [])
        if isinstance(r, dict) and str(r.get("client", "")).strip()
    }
    registered = set()
    try:
        import client_registry
        registered = {c["name"].strip() for c in client_registry.load_registry() if c.get("name")}
    except Exception as e:
        # Filter still works off whatever has rows; no reason to fail
        # the whole dashboard because the registry is briefly unreadable.
        print(f"Client filter falling back to rows-only ({e})")

    options = present | registered
    if allowed is not None:
        options &= allowed
    return sorted(options)


@app.route("/api/dashboard-data")
@login_required
def api_dashboard_data():
    role = session["role"]
    raw = load_dashboard_data()
    raw, overlay_ok = _overlay_feedback(raw)
    # Client scoping runs BEFORE the role filter and on the server, so a
    # restricted session never receives another client's rows at all -
    # not hidden in the browser, absent from the response.
    raw = permissions.filter_payload_for_clients(raw, _visible_clients_for_session())
    filtered = filter_payload_for_role(raw, role)
    filtered = dict(filtered)
    visible = _client_filter_options(raw, _visible_clients_for_session())
    # "none" is the difference between "this client genuinely has no
    # vehicles today" and "your account was never assigned a client, so
    # you are seeing nothing and always will until someone fixes it".
    # Those look identical on screen otherwise, and the second one is a
    # support call every time.
    scoped = _visible_clients_for_session()
    access = "all" if scoped is None else ("scoped" if scoped else "none")
    filtered["meta"] = {**filtered.get("meta", {}), "feedbackApplied": overlay_ok,
                        "clients": visible, "seesAllClients": permissions.sees_all_clients(role),
                        "clientAccess": access}
    # xlsxB64/tamperB64 never need to reach the browser at all, any role,
    # since /api/export/<name> reads them straight from disk on demand.
    filtered = {k: v for k, v in filtered.items() if k not in ("xlsxB64", "tamperB64")}
    return jsonify(filtered)


@app.route("/api/export/<name>")
@login_required
def api_export(name):
    role = session["role"]
    key_map = {"integrity": "integrity_xlsx", "tampering": "tampering_xlsx"}
    perm_key = key_map.get(name)
    if perm_key is None or role not in EXPORT_ACCESS.get(perm_key, ()):
        return jsonify({"error": "not permitted for your role"}), 403

    raw = load_dashboard_data()
    b64_key = "xlsxB64" if name == "integrity" else "tamperB64"
    b64 = raw.get(b64_key)
    if not b64:
        return jsonify({"error": "report not available"}), 404

    data = base64.b64decode(b64)
    # The workbook is generated fleet-wide, so downloading it bypassed
    # every other client boundary in the app - a GTL-only technician was
    # getting a spreadsheet containing 114 AGL vehicles. Filtered per
    # request rather than pre-generated per client: clients change, and
    # a stale per-client file would be worse than none.
    allowed = _visible_clients_for_session()
    if allowed is not None:
        try:
            data = _client_scoped_workbook(data, allowed)
        except Exception as e:
            # Never fall back to the unfiltered file - that is precisely
            # the leak. Fail the download instead.
            print(f"Could not scope the {name} export: {e}")
            return jsonify({"error": "Could not prepare a report limited to your clients."}), 503

    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=GTL_{name}_report.xlsx"},
    )


def _client_scoped_workbook(data, allowed):
    """
    Strips every row for a vehicle outside `allowed` from an already
    generated workbook.

    Filtering the finished file, rather than regenerating it from
    filtered data, keeps the formatting and formulas the standalone
    tool produces exactly as they are - the Executive Dashboard sheet
    is built from live COUNTA/COUNTIF formulas over the Full Data
    sheet, so removing rows there makes its totals correct on open
    without recomputing anything here.

    Any sheet whose header has no plate column is left alone; those are
    the settings/legend sheets, which carry no per-vehicle data.
    """
    import io
    from openpyxl import load_workbook

    owner = _plate_client_map()
    wb = load_workbook(io.BytesIO(data))
    for ws in wb.worksheets:
        plate_col = header_row = None
        # The header isn't always row 1 (some sheets open with a title
        # block), so scan the first few rows for the plate column.
        for r in range(1, min(ws.max_row, 8) + 1):
            for c in range(1, min(ws.max_column, 20) + 1):
                value = ws.cell(row=r, column=c).value
                if isinstance(value, str) and "plate" in value.strip().lower():
                    plate_col, header_row = c, r
                    break
            if plate_col:
                break
        if not plate_col:
            continue
        # Bottom-up so deleting one row doesn't shift the rows still to
        # be checked.
        for r in range(ws.max_row, header_row, -1):
            plate = ws.cell(row=r, column=plate_col).value
            if plate is None or str(plate).strip() == "":
                continue
            if owner.get(str(plate).strip(), "") not in allowed:
                ws.delete_rows(r)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.route("/api/feedback", methods=["POST"])
@login_required
def api_feedback():
    """
    All three roles can submit: this is a two-way log between whoever's
    watching the vehicle (client) and whoever acts on it (technician/
    admin), not a one-way form. requires_followup is an explicit choice
    made here, not guessed from wording - "No" is what pulls an asset
    out of the active Critical/Pending queues into Known Issues (see
    classifier.py), so it has to be a deliberate answer, not inferred.
    """
    role = session["role"]
    body = request.get_json(force=True, silent=True) or {}
    plate = (body.get("plate") or "").strip()
    comment = (body.get("comment") or "").strip()
    reported_by = (body.get("reportedBy") or "").strip()
    requires_followup_raw = body.get("requiresFollowup")
    entry_type = (body.get("entryType") or "feedback").strip().lower()

    if not plate or not comment or not reported_by:
        return jsonify({"error": "'plate', 'comment', and 'reportedBy' are required"}), 400
    if not isinstance(requires_followup_raw, bool):
        return jsonify({"error": "'requiresFollowup' must be true or false, an explicit choice, not optional"}), 400
    if entry_type not in ("feedback", "action"):
        return jsonify({"error": "'entryType' must be 'feedback' or 'action'"}), 400
    # The plate arrives from the request body, so without this a scoped
    # session can write feedback onto any client's vehicle - and that
    # write then triggers an email to that vehicle's real client.
    if not _plate_allowed(plate):
        return _deny_plate(plate)
    # Recommended Action is the technician's operational instruction to
    # the field - a client stating their own next step would be writing
    # your team's job card for them, so that field stays internal.
    if entry_type == "action" and role == "client":
        return jsonify({"error": "Only technicians and admins can set the Recommended Action"}), 403

    try:
        import sheets_store
        sheets_store.add_feedback(
            plate, comment, added_by=reported_by,
            requires_followup=requires_followup_raw, role=role,
            entry_type=entry_type,
        )
        status = sheets_store.infer_status(comment, requires_followup_raw)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        # Anything from gspread/Google's side (rate limit, timeout,
        # transient API error) landed here before as an unhandled
        # exception - Flask's HTML error page isn't valid JSON, so the
        # frontend's fetch().then(r => r.json()) throws and shows a
        # misleading "could not reach the server" for what was actually
        # a real, specific failure that reached this code just fine.
        return jsonify({"error": f"Could not save to Google Sheets right now: {e}"}), 503

    # The comment is already safely saved above - everything from here is
    # best-effort AND off the request thread. SMTP itself can take several
    # seconds (a blocked port timing out before mailer.py's fallback picks
    # up, or just a slow mail host), and there is nothing in that send the
    # browser needs before it can show "Saved" - the sheet write already
    # succeeded. Blocking the response on it turned every submit into a
    # multi-second wait for something the user was never looking at.
    base_url = public_base_url()
    respond_urls = None
    if role in ("admin", "technician"):
        respond_urls = {
            "no_followup": base_url + url_for("respond_page",
                token=make_respond_token(plate, "no_followup")),
            "needs_attention": base_url + url_for("respond_page",
                token=make_respond_token(plate, "needs_attention")),
        }
    _notify_async(plate, comment, reported_by, role, entry_type, requires_followup_raw, respond_urls)

    return jsonify({"ok": True, "plate": plate, "status": status})


@app.route("/feedback/respond")
def respond_page():
    """
    The page an email button lands on. Rendering it is a GET and changes
    nothing - the pre-selected answer only takes effect once the visitor
    reviews it and presses Confirm, which is a POST. This is deliberate:
    corporate mail scanners (Safe Links, Mimecast, Proofpoint...) prefetch
    every link in an email to scan it, and a GET that acted immediately
    would let a scanner silently answer on the client's behalf before
    they ever opened the message.
    """
    token = request.args.get("token", "")
    payload, error = read_respond_token(token)
    if error:
        return render_template("respond.html", error=error), 400
    try:
        import sheets_store
        if sheets_store.is_token_used(token):
            return render_template(
                "respond.html", error="This link has already been used to respond. "
                "If you need to add anything else, use the in-app comment form or contact us directly."), 400
    except Exception:
        pass  # can't check right now - fail open on the read, POST still enforces it authoritatively
    return render_template(
        "respond.html", plate=payload["plate"], action=payload["action"], token=token, error=None,
    )


@app.route("/feedback/respond", methods=["POST"])
def respond_submit():
    """
    Commits the answer from the confirm page. No login: the signed,
    expiring token IS the credential, scoped to exactly one plate and
    one pre-chosen action - it can set feedback for that vehicle and
    nothing else, no session, no fleet data, no other vehicle.
    """
    token = request.form.get("token", "") or (request.get_json(silent=True) or {}).get("token", "")
    payload, error = read_respond_token(token)
    if error:
        return jsonify({"error": error}), 400

    try:
        import sheets_store
        if sheets_store.is_token_used(token):
            return jsonify({"error": "This link has already been used to respond."}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503

    # The token proves WHICH vehicle this link may answer for - that's the
    # actual security boundary. The pre-selected action is only a default;
    # the visitor can still switch it on the page before confirming, the
    # same free choice the in-app form gives a logged-in user.
    plate = payload["plate"]
    body = request.get_json(silent=True) or request.form
    comment = (body.get("comment") or "").strip()
    name = (body.get("name") or "").strip()
    requires_followup_raw = body.get("requiresFollowup")
    if isinstance(requires_followup_raw, str):
        requires_followup_raw = requires_followup_raw.lower() == "true"

    if not name:
        return jsonify({"error": "Please tell us who's responding."}), 400
    if not isinstance(requires_followup_raw, bool):
        return jsonify({"error": "Please choose whether this needs follow-up."}), 400
    if not comment:
        comment = "No follow-up needed." if not requires_followup_raw else "Please look into this."

    requires_followup = requires_followup_raw
    try:
        import sheets_store
        sheets_store.add_feedback(
            plate, comment, added_by=name, requires_followup=requires_followup,
            role="client", entry_type="feedback",
        )
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Could not save your response right now: {e}"}), 503

    # Only mark the token spent once the response is actually saved - a
    # failed save above must leave the link usable for a retry, not
    # burn it on an attempt that never took effect.
    try:
        sheets_store.mark_token_used(token)
    except Exception as e:
        print(f"Could not record token as used for {plate}: {e}")

    _notify_async(plate, comment, name, "client", "feedback", requires_followup, None)

    return jsonify({"ok": True, "plate": plate})


@app.route("/api/feedback-history/<plate>")
@login_required
def api_feedback_history(plate):
    """
    The full trail for one vehicle, oldest first - what the "Customer
    Feedback" column on every table can't show (just the latest). Any
    logged-in role can read this: it's the same feedback the client
    themselves can already submit, not privileged data.

    Uses the cached read (same as /api/dashboard-data's overlay) rather
    than a fresh Sheets API call - the drill-down modal was paying a
    full ~1-4s Google Sheets round trip on every open, when the same
    30s-TTL cache everything else already relies on is fresh enough
    here too. A submit on this exact plate still lands instantly for
    the submitter: add_feedback() patches this worker's cache in place
    before this ever runs (see sheets_store._patch_cache_with).
    """
    # The plate is a URL path segment, so this endpoint was a direct
    # read of any client's comment trail regardless of what the
    # dashboard showed.
    if not _plate_allowed(plate):
        return _deny_plate(plate)
    try:
        import sheets_store
        all_feedback = sheets_store.load_feedback_cached()
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Could not read from Google Sheets right now: {e}"}), 503

    import sys
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "fleet_logic"))
    from schema import normalize_plate
    entry = all_feedback.get(normalize_plate(plate))
    history = entry["history"] if entry else []
    return jsonify({
        "plate": plate,
        "history": [
            {
                "comment": h["comment"], "status": h["status"], "requiresFollowup": h["requiresFollowup"],
                "date": h["date"].strftime("%d %b %Y, %H:%M") if h["date"].year > 1 else "",
                "addedBy": h["addedBy"], "role": h["role"], "entryType": h.get("entryType", "feedback"),
            }
            for h in history
        ],
    })


@app.route("/api/feedback-activity")
@login_required
def api_feedback_activity():
    """
    The global activity feed: every comment across every vehicle,
    newest first, so anyone with the dashboard open can see what's been
    reported today (or any other day) without opening each vehicle one
    at a time. Same data class as api_feedback_history, every role can
    read this.
    """
    try:
        import sheets_store
        entries = sheets_store.load_all_feedback_entries()
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Could not read from Google Sheets right now: {e}"}), 503

    # dateISO (YYYY-MM-DD) is what the frontend actually filters/counts
    # on - unambiguous and locale-independent, unlike matching "26 Jul
    # 2026"-style strings between Python and JS, which breaks the
    # moment either side's month abbreviation or locale differs.
    # "date"/"dateOnly" stay as display-formatted strings for showing
    # in the UI, never for comparison.
    # Comments are stored per plate with no client of their own, so they
    # need the same plate-based scoping the tampering sections get -
    # otherwise this endpoint hands a GTL-scoped user every AGL comment,
    # neatly bypassing the filtering on /api/dashboard-data.
    owner = _plate_client_map()
    allowed = _visible_clients_for_session()
    if allowed is not None:
        entries = [e for e in entries if owner.get(str(e.get("plate", "")).strip(), "") in allowed]

    today_iso = now_eat().strftime("%Y-%m-%d")
    serialized = [
        {
            "plate": e["plate"], "comment": e["comment"], "requiresFollowup": e["requiresFollowup"],
            "date": e["date"].strftime("%d %b %Y, %H:%M") if e["date"].year > 1 else "",
            "dateOnly": e["date"].strftime("%d %b %Y") if e["date"].year > 1 else "",
            "dateISO": e["date"].strftime("%Y-%m-%d") if e["date"].year > 1 else "",
            "addedBy": e["addedBy"], "role": e["role"], "entryType": e.get("entryType", "feedback"),
            "client": owner.get(str(e.get("plate", "")).strip(), ""),
        }
        for e in entries
    ]
    today_count = sum(1 for e in serialized if e["dateISO"] == today_iso)

    return jsonify({"todayCount": today_count, "todayISO": today_iso, "entries": serialized})


REFRESH_LOCK_PATH = os.path.join(os.path.dirname(__file__), "data", "_refresh_lock.json")
STALE_THRESHOLD_HOURS = 2
LOCK_COOLDOWN_MINUTES = 10


def _claim_refresh_lock():
    """
    True if this call just claimed the right to trigger a refresh,
    False if someone else already holds it. Uses O_CREAT|O_EXCL, which
    is atomic at the OS level, so this is safe across gunicorn's
    multiple worker *processes* on Render, not just threads in one -
    a plain in-memory flag would only ever be seen by one worker.
    """
    try:
        fd = os.open(REFRESH_LOCK_PATH, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w") as f:
            json.dump({"triggered_at": time.time()}, f)
        return True
    except FileExistsError:
        try:
            with open(REFRESH_LOCK_PATH) as f:
                held = json.load(f)
            age_minutes = (time.time() - held.get("triggered_at", 0)) / 60
        except (OSError, ValueError, json.JSONDecodeError):
            age_minutes = LOCK_COOLDOWN_MINUTES  # unreadable lock, treat as stale, safe to steal
        if age_minutes < LOCK_COOLDOWN_MINUTES:
            return False
        try:
            os.remove(REFRESH_LOCK_PATH)
        except OSError:
            pass
        return _claim_refresh_lock()  # one retry now that the stale lock is cleared


@app.route("/api/refresh-if-stale", methods=["POST"])
@login_required
def api_refresh_if_stale():
    """
    Backup for when the Apps Script scheduler doesn't fire. Any logged-
    in user's dashboard load can call this, but staleness and the
    trigger lock are both decided here, server-side - never trusted
    from the client (clocks lie, and coordinating "who goes first"
    across browser tabs from different users isn't something the
    client can do safely anyway). At most one real import gets kicked
    off per LOCK_COOLDOWN_MINUTES no matter how many requests hit this
    at once.
    """
    try:
        with open(DATA_PATH) as f:
            meta = json.load(f).get("meta", {})
        # importedAt (wall-clock time the import last actually ran) is
        # the right signal here, not "generated" (the latest timestamp
        # found IN the report data, which lags real time by design -
        # see run_import.py). Older data written before this field
        # existed won't have it; treat that as stale too, correctly,
        # since it means no import has run since this feature shipped.
        imported_str = meta.get("importedAt", "")
        imported_at = datetime.strptime(imported_str, "%d %B %Y, %H:%M")
        age_hours = (now_eat() - imported_at).total_seconds() / 3600
    except (FileNotFoundError, ValueError, KeyError):
        age_hours = STALE_THRESHOLD_HOURS + 1  # no readable/importedAt-less data - treat as stale

    if age_hours < STALE_THRESHOLD_HOURS:
        return jsonify({"status": "fresh", "ageHours": round(age_hours, 1)})

    if not _claim_refresh_lock():
        return jsonify({"status": "refresh_recently_triggered", "ageHours": round(age_hours, 1)})

    def _background_import():
        try:
            from run_import import run_import as do_import
            do_import(force=True)
        except Exception as e:
            print(f"Background auto-refresh failed: {e}")

    threading.Thread(target=_background_import, daemon=True).start()
    return jsonify({"status": "refresh_triggered", "ageHours": round(age_hours, 1)})


@app.route("/api/checkin/trigger", methods=["POST"])
@login_required
def api_checkin_trigger():
    """
    The dashboard's "Send check-in now" button (admin/technician only -
    this pulls fresh mail and sends real email to the client, not
    something to expose to a client-role session). Runs a real import
    (force=True, same as the stale-refresh path above) so the digests
    reflect current data, then force-sends the pending-confirmation,
    technical-escalation, and known-issues-checkin digests regardless
    of their normal weekly interval - see run_import.py's force_digests
    param. The tampering report is NOT part of this button; it stays on
    its own schedule (see notifications.send_tamper_risk_report_digest).

    Runs in the background (a real IMAP fetch + Sheets reads/writes can
    take a while) - the caller finds out who last triggered it via
    GET /api/checkin/last below, not from this response.
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    username = session["username"]

    def _background_checkin():
        try:
            from run_import import run_import as do_import
            do_import(force=True, force_digests=True, prompted_by=username)
        except Exception as e:
            print(f"Manual check-in failed: {e}")

    threading.Thread(target=_background_checkin, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/checkin/last", methods=["GET"])
@login_required
def api_checkin_last():
    """Who last pressed the button, and when - so everyone looking at
    the dashboard sees the same answer, not just whoever clicked it."""
    try:
        import sheets_store
        last = sheets_store.get_last_manual_checkin()
    except Exception:
        last = None
    if not last:
        return jsonify({"lastCheckin": None})
    return jsonify({"lastCheckin": {"by": last["by"], "at": last["at"].strftime("%d %b %Y, %H:%M")}})


@app.route("/api/tamper-check", methods=["POST"])
@login_required
def api_tamper_check():
    """
    Records that staff physically inspected a vehicle flagged by the
    tampering detector. See sheets_store.add_tamper_check()'s docstring:
    from this point on, any gap on this plate dated on or before now is
    excluded from confirmed/unconfirmed on the NEXT import (this
    doesn't retroactively edit the currently-loaded dashboard data).
    Admin/technician only - a client isn't the one doing the physical
    check.
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    body = request.get_json(force=True, silent=True) or {}
    plate = (body.get("plate") or "").strip()
    comment = (body.get("comment") or "").strip()
    if not plate or not comment:
        return jsonify({"error": "'plate' and 'comment' are required"}), 400
    # This write suppresses future tampering cases for the vehicle, so
    # an unscoped version let one client's technician silently mute
    # another client's tamper alerts.
    if not _plate_allowed(plate):
        return _deny_plate(plate)

    try:
        import sheets_store
        sheets_store.add_tamper_check(plate, session["username"], comment)
    except Exception as e:
        return jsonify({"error": f"Could not save the check: {e}"}), 503
    return jsonify({"ok": True, "plate": plate})


def _write_json_atomic(path, data):
    """Poll snapshots are read by refresh_live_snapshot() (fired right
    after every successful poll) while the next poll may be writing
    them - see fleet_logic/atomic_json.py for why that needs more care
    than open()+dump()."""
    atomic_json.write_json_atomic(path, data, indent=2)


MIX_API_SNAPSHOT_PATH = os.path.join(os.path.dirname(__file__), "data", "mix_api_snapshot.json")


def _load_settings():
    return _load_settings_from(os.path.join(os.path.dirname(__file__), "data", "settings.ini"))


def _load_settings_from(path):
    from settings import load_settings
    return load_settings(path)


def _poll_registry(settings):
    """The client list every poller works from. See
    client_registry.load_registry() for the Sheets -> cache ->
    settings.ini fallback chain and why it never raises."""
    import client_registry
    return client_registry.load_registry(settings)


def _mix_api_poll_once():
    """
    One fetch-assets-and-positions cycle across every MiX organisation
    claimed by a client in the registry, written to
    MIX_API_SNAPSHOT_PATH. Rows are tagged with the CANONICAL CLIENT
    NAME, not just the raw org id, so everything downstream (access
    control, the dashboard's client filter) can group by client without
    knowing anything about MiX's own ids.
    """
    import client_registry
    from adapters.mix_api_client import MixApiClient
    from adapters import mix_api

    settings = _load_settings()
    clients = _poll_registry(settings)
    org_ids = client_registry.ids_for_platform(clients, "mix")
    if not org_ids:
        return {"status": "skipped", "reason": "no MiX organisations mapped to any client"}
    owner = client_registry.platform_index(clients)["mix"]

    client = MixApiClient(inter_org_delay_seconds=settings["MIX_API_INTER_ORG_DELAY_SECONDS"])
    if not client.is_configured():
        return {"status": "skipped", "reason": "MIX_CLIENT_ID/MIX_CLIENT_SECRET/MIX_USERNAME/MIX_PASSWORD not set"}

    reports, retired_plates = mix_api.fetch_all_reports_and_decommissioned(client, org_ids)
    by_org = {}
    for r in reports:
        by_org.setdefault(r.organisation_id, []).append({
            "plate": r.asset_plate,
            "client": owner.get(str(r.organisation_id), client_registry.UNASSIGNED),
            "lat": r.last_lat,
            "lon": r.last_lon,
            "locationText": r.last_location_text,
            "lastReportTime": r.last_report_time.isoformat() if r.last_report_time else None,
            "assetId": r.raw_row["asset"].get("AssetId"),
            "fleetNumber": r.raw_row["asset"].get("FleetNumber"),
        })
    snapshot = {
        "fetchedAt": now_eat().isoformat(),
        "orgIds": org_ids,
        "clientByOrgId": owner,
        # Applied by run_import.process_reports() to EVERY source, not
        # just MiX - see mix_api.decommissioned_plates() for why the
        # mailed reports would otherwise reintroduce retired vehicles.
        "decommissionedPlates": sorted(retired_plates),
        "counts": {org: len(rows) for org, rows in by_org.items()},
        "byOrg": by_org,
    }
    _write_json_atomic(MIX_API_SNAPSHOT_PATH, snapshot)
    return {"status": "ok", "counts": snapshot["counts"]}


_last_live_refresh = {"at": 0.0}
_live_refresh_lock = threading.Lock()
# All three pollers share a cadence, so they finish within seconds of
# each other and each used to trigger a full reclassification - three
# runs producing near-identical output, each re-reading feedback,
# tamper checks and vehicle status from Sheets. That tripled the work
# and the Sheets traffic for no added freshness, and contributed to
# hitting the Sheets read-per-minute quota. One refresh per window is
# enough: whichever poller finishes first picks up the other two
# platforms' snapshots anyway, since they're read from disk.
_LIVE_REFRESH_MIN_INTERVAL_SECONDS = 60


def _refresh_live_snapshot_after_poll(platform_label):
    """
    Shared tail call for every platform's poll loop: once a live API
    poll has fresh data on disk, immediately reclassify the dashboard
    from it (run_import.refresh_live_snapshot()) instead of waiting for
    tomorrow's mail-based run_import() - this is what makes the
    dashboard's "Updated ..." pill actually track the ~5 min API
    cadence rather than the once-daily import. See that function's own
    docstring for why it's safe to call this often (no notifications,
    no day-over-day status persistence, meta.importedAt untouched).
    """
    with _live_refresh_lock:
        since = time.time() - _last_live_refresh["at"]
        if since < _LIVE_REFRESH_MIN_INTERVAL_SECONDS:
            print(f"Live snapshot refresh skipped after {platform_label} poll: "
                  f"another ran {since:.0f}s ago")
            return
        _last_live_refresh["at"] = time.time()

    try:
        from run_import import refresh_live_snapshot
        result = refresh_live_snapshot()
        if result["status"] == "ok":
            print(f"Live snapshot refreshed after {platform_label} poll: generated {result['generated']}")
        elif result["status"] == "skipped":
            print(f"Live snapshot refresh skipped after {platform_label} poll: {result['reason']}")
        else:
            print(f"Live snapshot refresh failed after {platform_label} poll: {result.get('reason')}")
    except Exception as e:
        print(f"Live snapshot refresh errored after {platform_label} poll: {e}")


def _mix_api_poll_loop():
    while True:
        try:
            result = _mix_api_poll_once()
            if result["status"] == "ok":
                print(f"MiX API poll: {result['counts']}")
                _refresh_live_snapshot_after_poll("MiX")
            else:
                print(f"MiX API poll skipped: {result['reason']}")
        except Exception as e:
            print(f"MiX API poll failed: {e}")

        from settings import load_settings
        settings = load_settings(os.path.join(os.path.dirname(__file__), "data", "settings.ini"))
        time.sleep(max(1, settings["MIX_API_POLL_INTERVAL_MINUTES"]) * 60)


def _start_mix_api_poller():
    """
    Fires once at app startup and keeps running for as long as this
    process does (see the use_reloader=False note at the bottom of this
    file for why local dev doesn't double it up). No-op if the process
    dies/restarts (Heroku dyno cycling, deploys, crashes) - there's no
    persistence across that, same tradeoff as every other "while the
    app happens to be running" timer in this codebase.
    """
    threading.Thread(target=_mix_api_poll_loop, daemon=True).start()


def _mix_api_poller_enabled():
    """
    Guards against starting a real background network poller as a side
    effect of merely importing app.py - test_permissions.py does exactly
    that (from app import app). Checking sys.modules for "pytest" (set
    from the moment pytest starts, unlike PYTEST_CURRENT_TEST which
    only exists during an individual test's execution, not at import/
    collection time) keeps the test suite from making live MiX API
    calls. DISABLE_MIX_API_POLLER is the manual escape hatch for anyone
    else importing this module without meaning to run the real server.
    """
    if "pytest" in sys.modules:
        return False
    if os.environ.get("DISABLE_MIX_API_POLLER") == "true":
        return False
    return True


if _mix_api_poller_enabled():
    _start_mix_api_poller()


@app.route("/api/mix/snapshot", methods=["GET"])
@login_required
def api_mix_snapshot():
    """Read-only view of the last MiX API poll, for verifying the
    ingestion path."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    return _scoped_snapshot(MIX_API_SNAPSHOT_PATH, "byOrg")


@app.route("/api/mix/poll-now", methods=["POST"])
@login_required
def api_mix_poll_now():
    """Manual trigger so testing doesn't require waiting for the
    interval - runs in the background, check /api/mix/snapshot after."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    threading.Thread(target=_mix_api_poll_once, daemon=True).start()
    return jsonify({"status": "started"})


TELETRAC_API_SNAPSHOT_PATH = os.path.join(os.path.dirname(__file__), "data", "teletrac_api_snapshot.json")


def _teletrac_api_poll_once():
    """
    One fetch-current-data-for-every-configured-client cycle, written
    to TELETRAC_API_SNAPSHOT_PATH. Same shape as _mix_api_poll_once
    above, mirrored for Teletrac's Integrate API (see
    fleet_logic/adapters/teletrac_api_client.py).
    """
    import client_registry
    from adapters.teletrac_api_client import TeletracApiClient
    from adapters import teletrac_api

    settings = _load_settings()
    clients = _poll_registry(settings)
    client_ids = client_registry.ids_for_platform(clients, "teletrac")
    if not client_ids:
        return {"status": "skipped", "reason": "no Teletrac clients mapped to any client"}
    owner = client_registry.platform_index(clients)["teletrac"]

    client = TeletracApiClient(inter_client_delay_seconds=settings["TELETRAC_API_INTER_CLIENT_DELAY_SECONDS"])
    if not client.is_configured():
        return {"status": "skipped", "reason": "TELETRAC_API_KEY/TELETRAC_API_SECRET not set"}

    reports = teletrac_api.fetch_all_reports(client, client_ids)
    by_client = {}
    for r in reports:
        by_client.setdefault(r.organisation_id, []).append({
            "plate": r.asset_plate,
            "client": owner.get(str(r.organisation_id), client_registry.UNASSIGNED),
            "lat": r.last_lat,
            "lon": r.last_lon,
            "locationText": r.last_location_text,
            "lastReportTime": r.last_report_time.isoformat() if r.last_report_time else None,
            "imei": r.imei,
        })
    snapshot = {
        "fetchedAt": now_eat().isoformat(),
        "clientIds": client_ids,
        "clientByClientId": owner,
        "counts": {c: len(rows) for c, rows in by_client.items()},
        "byClient": by_client,
    }
    _write_json_atomic(TELETRAC_API_SNAPSHOT_PATH, snapshot)
    return {"status": "ok", "counts": snapshot["counts"]}


def _teletrac_api_poll_loop():
    while True:
        try:
            result = _teletrac_api_poll_once()
            if result["status"] == "ok":
                print(f"Teletrac API poll: {result['counts']}")
                _refresh_live_snapshot_after_poll("Teletrac")
            else:
                print(f"Teletrac API poll skipped: {result['reason']}")
        except Exception as e:
            print(f"Teletrac API poll failed: {e}")

        from settings import load_settings
        settings = load_settings(os.path.join(os.path.dirname(__file__), "data", "settings.ini"))
        time.sleep(max(1, settings["TELETRAC_API_POLL_INTERVAL_MINUTES"]) * 60)


def _start_teletrac_api_poller():
    """Fires once at app startup and keeps running for as long as this
    process does - same lifetime/restart tradeoffs as _start_mix_api_poller."""
    threading.Thread(target=_teletrac_api_poll_loop, daemon=True).start()


def _teletrac_api_poller_enabled():
    """Same guard as _mix_api_poller_enabled - keeps importing app.py
    (e.g. test_permissions.py) from starting a real background poller."""
    if "pytest" in sys.modules:
        return False
    if os.environ.get("DISABLE_TELETRAC_API_POLLER") == "true":
        return False
    return True


if _teletrac_api_poller_enabled():
    _start_teletrac_api_poller()


@app.route("/api/teletrac/snapshot", methods=["GET"])
@login_required
def api_teletrac_snapshot():
    """Read-only view of the last Teletrac API poll, for verifying the
    new ingestion path before/after it's wired into the real dashboard data."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    return _scoped_snapshot(TELETRAC_API_SNAPSHOT_PATH, "byClient")


@app.route("/api/teletrac/poll-now", methods=["POST"])
@login_required
def api_teletrac_poll_now():
    """Manual trigger so testing doesn't require waiting for the
    interval - runs in the background, check /api/teletrac/snapshot after."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    threading.Thread(target=_teletrac_api_poll_once, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/teletrac/clients", methods=["GET"])
@login_required
def api_teletrac_clients():
    """Lists every client visible to these Teletrac API credentials, so
    an operator can look up the ClientID for 'Globe Trotters Ltd' to put
    in settings.ini's [teletrac_api] client_ids - there's no other way
    to discover it short of asking Teletrac support."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    from adapters.teletrac_api_client import TeletracApiClient
    client = TeletracApiClient()
    if not client.is_configured():
        return jsonify({"error": "TELETRAC_API_KEY/TELETRAC_API_SECRET not set"}), 503
    try:
        return jsonify({"clients": client.get_all_clients()})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


FT_CLOUD_API_SNAPSHOT_PATH = os.path.join(os.path.dirname(__file__), "data", "ft_cloud_api_snapshot.json")


def _ft_cloud_api_poll_once():
    """
    One fetch-vehicles-and-device-status-for-every-configured-fleet
    cycle, written to FT_CLOUD_API_SNAPSHOT_PATH. Same shape as
    _mix_api_poll_once/_teletrac_api_poll_once above, mirrored for FT
    Cloud's OpenAPI (see fleet_logic/adapters/ft_cloud_api_client.py).
    """
    import client_registry
    from adapters.ft_cloud_api_client import FtCloudApiClient
    from adapters import ft_cloud_api

    settings = _load_settings()
    clients = _poll_registry(settings)
    fleet_ids = client_registry.ids_for_platform(clients, "ftCloud")
    if not fleet_ids:
        return {"status": "skipped", "reason": "no FT Cloud fleets mapped to any client"}
    owner = client_registry.platform_index(clients)["ftCloud"]

    client = FtCloudApiClient(inter_call_delay_seconds=settings["FT_CLOUD_API_INTER_CALL_DELAY_SECONDS"])
    if not client.is_configured():
        return {"status": "skipped", "reason": "FT_CLOUD_API_SIGN/FT_CLOUD_TENANT_ID not set"}

    # Coordinates come from one of three places - see settings.ini's
    # [ft_cloud_api] position_source. "webhook" is the cheap path: the
    # positions were already pushed to this app by FT, so the poll
    # makes ZERO extra calls for them instead of one per device.
    source = settings["FT_CLOUD_API_POSITION_SOURCE"]
    webhook_positions = None
    webhook_last_seen = None
    if source == "webhook":
        from adapters import ft_cloud_webhook
        webhook_positions = ft_cloud_webhook.positions_by_unique_id(
            FT_CLOUD_WEBHOOK_STATE_PATH,
            max_age_minutes=settings["FT_CLOUD_API_WEBHOOK_POSITION_MAX_AGE_MINUTES"])
        # NOT freshness-filtered on purpose - a device whose last webhook
        # report was over a week ago must still show that stale
        # timestamp for online/offline purposes, not fall back to FT's
        # connectivity flag (which stayed "online" the whole time). See
        # ft_cloud_api._last_seen().
        webhook_last_seen = ft_cloud_webhook.last_seen_by_unique_id(FT_CLOUD_WEBHOOK_STATE_PATH)

    reports = ft_cloud_api.fetch_all_reports(
        client, fleet_ids,
        fetch_positions=(source == "trips"),
        position_lookback_days=settings["FT_CLOUD_API_POSITION_LOOKBACK_DAYS"],
        preset_positions=webhook_positions,
        webhook_last_seen=webhook_last_seen)
    by_fleet = {}
    for r in reports:
        by_fleet.setdefault(r.organisation_id, []).append({
            "plate": r.asset_plate,
            "client": owner.get(str(r.organisation_id), client_registry.UNASSIGNED),
            "lat": r.last_lat,
            "lon": r.last_lon,
            "lastReportTime": r.last_report_time.isoformat() if r.last_report_time else None,
            "statusNote": r.status_note,
            "uniqueId": r.raw_row["deviceInfo"].get("uniqueId"),
        })
    snapshot = {
        "fetchedAt": now_eat().isoformat(),
        "fleetIds": fleet_ids,
        "clientByFleetId": owner,
        "positionSource": source,
        "counts": {f: len(rows) for f, rows in by_fleet.items()},
        "byFleet": by_fleet,
    }
    _write_json_atomic(FT_CLOUD_API_SNAPSHOT_PATH, snapshot)
    return {"status": "ok", "counts": snapshot["counts"], "positionSource": source,
            "withCoordinates": sum(1 for rows in by_fleet.values() for r in rows if r["lat"] is not None)}


def _ft_cloud_api_poll_loop():
    while True:
        try:
            result = _ft_cloud_api_poll_once()
            if result["status"] == "ok":
                print(f"FT Cloud API poll: {result['counts']}")
                _refresh_live_snapshot_after_poll("FT Cloud")
            else:
                print(f"FT Cloud API poll skipped: {result['reason']}")
        except Exception as e:
            print(f"FT Cloud API poll failed: {e}")

        from settings import load_settings
        settings = load_settings(os.path.join(os.path.dirname(__file__), "data", "settings.ini"))
        time.sleep(max(1, settings["FT_CLOUD_API_POLL_INTERVAL_MINUTES"]) * 60)


def _start_ft_cloud_api_poller():
    """Fires once at app startup and keeps running for as long as this
    process does - same lifetime/restart tradeoffs as _start_mix_api_poller."""
    threading.Thread(target=_ft_cloud_api_poll_loop, daemon=True).start()


def _ft_cloud_api_poller_enabled():
    """Same guard as _mix_api_poller_enabled - keeps importing app.py
    (e.g. test_permissions.py) from starting a real background poller."""
    if "pytest" in sys.modules:
        return False
    if os.environ.get("DISABLE_FT_CLOUD_API_POLLER") == "true":
        return False
    return True


if _ft_cloud_api_poller_enabled():
    _start_ft_cloud_api_poller()


@app.route("/api/ftcloud/snapshot", methods=["GET"])
@login_required
def api_ft_cloud_snapshot():
    """Read-only view of the last FT Cloud API poll, for verifying the
    new ingestion path before/after it's wired into the real dashboard data."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    return _scoped_snapshot(FT_CLOUD_API_SNAPSHOT_PATH, "byFleet")


@app.route("/api/ftcloud/poll-now", methods=["POST"])
@login_required
def api_ft_cloud_poll_now():
    """Manual trigger so testing doesn't require waiting for the
    interval - runs in the background, check /api/ftcloud/snapshot after."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    threading.Thread(target=_ft_cloud_api_poll_once, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/ftcloud/fleets", methods=["GET"])
@login_required
def api_ft_cloud_fleets():
    """Lists every fleet visible to these FT Cloud credentials, so an
    operator can look up the fleetId for 'Globe Trotters Ltd(GTL)' to
    put in settings.ini's [ft_cloud_api] fleet_ids."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    from adapters.ft_cloud_api_client import FtCloudApiClient
    client = FtCloudApiClient()
    if not client.is_configured():
        return jsonify({"error": "FT_CLOUD_API_SIGN/FT_CLOUD_TENANT_ID not set"}), 503
    try:
        return jsonify({"fleets": client.get_fleets()})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


FT_CLOUD_WEBHOOK_STATE_PATH = os.path.join(os.path.dirname(__file__), "data", "ft_cloud_webhook_state.json")


def _ft_cloud_webhook_secret():
    """
    The receiver route below is necessarily PUBLIC - FT posts to it
    unauthenticated, and its subscribe API accepts only a callbackUrl,
    with no way to attach a header or credential. So the only place a
    shared secret can live is the URL path itself, which makes that
    URL a bearer token: treat it like a password, and rotate it by
    changing this value and re-subscribing.
    """
    return os.environ.get("FT_CLOUD_WEBHOOK_SECRET") or ""


def _ft_cloud_subscribed_unique_ids():
    """
    The device allowlist for the receiver. Read from the last poll's
    snapshot rather than by calling FT, so a webhook delivery never
    triggers an outbound API call - this route is on FT's hot path and
    must stay fast. Returns None (meaning "allow anything") only when
    no snapshot exists yet, so a freshly deployed app still captures
    positions before its first poll completes.
    """
    if not os.path.exists(FT_CLOUD_API_SNAPSHOT_PATH):
        return None
    try:
        with open(FT_CLOUD_API_SNAPSHOT_PATH) as f:
            snapshot = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    ids = {r.get("uniqueId") for rows in snapshot.get("byFleet", {}).values()
           for r in rows if r.get("uniqueId")}
    return ids or None


@app.route("/webhook/ftcloud/<secret>", methods=["POST"])
def ft_cloud_webhook_receiver(secret):
    """
    Public receiver for FT's push stream. Deliberately NOT
    @login_required - FT is a machine posting to it, not a browser
    session. Access control is the secret in the path (see
    _ft_cloud_webhook_secret) plus the device allowlist.

    Always answers 200 once the secret checks out, even on a payload
    it could not parse: providers commonly disable a subscription
    after repeated non-200s, and losing the whole stream because of
    one malformed message would be a far worse failure than dropping
    that message. Anything unparsed is retained in the state file for
    inspection instead.
    """
    expected = _ft_cloud_webhook_secret()
    if not expected or not hmac.compare_digest(secret, expected):
        return jsonify({"error": "not found"}), 404

    from adapters import ft_cloud_webhook
    try:
        body = request.get_json(force=True, silent=True)
        result = ft_cloud_webhook.record_events(
            FT_CLOUD_WEBHOOK_STATE_PATH, body,
            message_type=request.args.get("type"),
            allowed_unique_ids=_ft_cloud_subscribed_unique_ids())
    except Exception as e:
        print(f"FT Cloud webhook receiver error (answering 200 anyway): {e}")
        return jsonify({"ok": True}), 200
    return jsonify({"ok": True, **result}), 200


@app.route("/api/ftcloud/webhook/subscribe", methods=["POST"])
@login_required
def api_ft_cloud_webhook_subscribe():
    """
    One-off registration of this app as FT's push target, scoped to
    the configured fleets' devices only.

    Two things this deliberately refuses rather than guesses at:
    a base URL it can't verify is publicly reachable, and an empty
    device list (which FT documents as "subscribe to all devices" -
    on this reseller tenant that would pull in 10 other companies'
    vehicles).
    """
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403

    secret = _ft_cloud_webhook_secret()
    if not secret:
        return jsonify({"error": "FT_CLOUD_WEBHOOK_SECRET is not set - refusing to expose an "
                                  "unauthenticated public webhook endpoint"}), 400
    base = public_base_url()
    if not base:
        return jsonify({"error": "No public base URL available (PUBLIC_BASE_URL / "
                                  "RENDER_EXTERNAL_URL). FT must be able to reach this app."}), 400

    from settings import load_settings
    from adapters.ft_cloud_api_client import FtCloudApiClient
    settings = load_settings(os.path.join(os.path.dirname(__file__), "data", "settings.ini"))
    fleet_ids = settings["FT_CLOUD_API_FLEET_IDS"]
    if not fleet_ids:
        return jsonify({"error": "no fleet_ids configured in settings.ini [ft_cloud_api]"}), 400

    client = FtCloudApiClient()
    if not client.is_configured():
        return jsonify({"error": "FT_CLOUD_API_SIGN/FT_CLOUD_TENANT_ID not set"}), 503

    types = request.json.get("types") if request.is_json and request.json else None
    types = types or ["GPS", "ONLINE_STATE"]
    callback_url = f"{base}/webhook/ftcloud/{secret}"

    try:
        unique_ids = client.get_fleet_unique_ids(fleet_ids)
    except Exception as e:
        return jsonify({"error": f"could not list fleet devices: {e}"}), 502

    results = {}
    for t in types:
        try:
            client.subscribe_webhook(t, f"{callback_url}?type={t}", unique_ids)
            results[t] = "subscribed"
        except Exception as e:
            results[t] = f"failed: {e}"
    return jsonify({"deviceCount": len(unique_ids), "callbackUrl": callback_url.replace(secret, "***"),
                     "results": results})


@app.route("/api/ftcloud/webhook/status", methods=["GET"])
@login_required
def api_ft_cloud_webhook_status():
    """What FT thinks is subscribed, plus what this app has actually
    received - the two together are what tell you whether the push
    stream is genuinely flowing or just registered."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    from adapters.ft_cloud_api_client import FtCloudApiClient
    from adapters import ft_cloud_webhook

    client = FtCloudApiClient()
    remote = {}
    if client.is_configured():
        for t in ("GPS", "ONLINE_STATE"):
            try:
                sub = client.get_webhook_subscription(t)
                remote[t] = {"subscribed": sub.get("subscribed"),
                              "deviceCount": len([u for u in (sub.get("uniqueIds") or "").split(",") if u])}
            except Exception as e:
                remote[t] = {"error": str(e)}

    state = ft_cloud_webhook.load_state(FT_CLOUD_WEBHOOK_STATE_PATH)
    return jsonify({
        "secretConfigured": bool(_ft_cloud_webhook_secret()),
        "publicBaseUrl": public_base_url() or None,
        "subscriptions": remote,
        "received": {
            "devices": len(state.get("devices", {})),
            "withCoordinates": sum(1 for d in state.get("devices", {}).values() if d.get("lat") is not None),
            "updatedAt": state.get("updatedAt"),
            "unparsedSamples": state.get("unparsed", [])[:3],
        },
    })


@app.route("/api/ftcloud/webhook/unsubscribe", methods=["POST"])
@login_required
def api_ft_cloud_webhook_unsubscribe():
    """Hands the push stream back. Useful if Teletrac needs the
    tenant's single callbackUrl-per-type for another integration."""
    if session["role"] not in MANAGE_USERS_ROLES:
        return jsonify({"error": "Not permitted for your role"}), 403
    from adapters.ft_cloud_api_client import FtCloudApiClient
    client = FtCloudApiClient()
    if not client.is_configured():
        return jsonify({"error": "FT_CLOUD_API_SIGN/FT_CLOUD_TENANT_ID not set"}), 503
    types = (request.json or {}).get("types") if request.is_json else None
    results = {}
    for t in (types or ["GPS", "ONLINE_STATE"]):
        try:
            client.unsubscribe_webhook(t)
            results[t] = "unsubscribed"
        except Exception as e:
            results[t] = f"failed: {e}"
    return jsonify({"results": results})


@app.route("/api/import", methods=["GET", "POST"])
def api_import():
    """
    Called ONLY by the Apps Script scheduler, never by a browser.
    Protected by a shared secret in the X-API-Key header, not a user
    session, since there's no logged-in user at 4:15 AM.
    """
    expected_key = os.environ.get("IMPORT_API_KEY")
    provided_key = request.headers.get("X-API-Key")
    if not expected_key or provided_key != expected_key:
        return jsonify({"error": "unauthorized"}), 403

    force = request.args.get("force") == "true"
    try:
        from run_import import run_import as do_import
        result = do_import(force=force)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/")
def index():
    if "username" not in session:
        return redirect(url_for("login_page"))
    role = session["role"]
    return render_template(
        "dashboard.html", allowed_panels=allowed_panels(role),
        can_export_integrity=role in EXPORT_ACCESS["integrity_xlsx"],
        can_export_tampering=role in EXPORT_ACCESS["tampering_xlsx"],
        can_manage_users=role in MANAGE_USERS_ROLES,
        # Built from the sheet ID rather than stored as a second env var,
        # so there's only ever one place the spreadsheet is identified.
        # Empty when Sheets isn't configured, and the button hides itself.
        sheet_url=(
            f"https://docs.google.com/spreadsheets/d/{os.environ['FEEDBACK_SHEET_ID']}/edit"
            if os.environ.get("FEEDBACK_SHEET_ID") else ""
        ),
    )


@app.route("/login")
def login_page():
    if "username" in session:
        return redirect(url_for("index"))
    return render_template("login.html")


if __name__ == "__main__":
    # use_reloader=False: the reloader re-imports this module in a
    # separate watcher process, which would start a second copy of the
    # MiX API poll thread below (see _start_mix_api_poller) racing the
    # real one - not harmful (idempotent snapshot writes), just double
    # API traffic for no reason during local dev.
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, port=port, use_reloader=False)
