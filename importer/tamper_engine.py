"""
Device Tampering Risk Report - core engine, ported verbatim.

This is the exact analyse() logic from tampering_report.py, the
standalone GUI/CLI tool. Every constant, every threshold, every
classification rule is unchanged. The only thing that's different is
where the input CSVs come from (Gmail import vs. manual file picker)
and where the output goes (JSON for the dashboard API instead of only
an .xlsx file) - the detection logic itself is untouched, byte-for-byte
the same decisions.

If this file's output ever needs to be re-verified against the
original tool, run test_tamper_engine.py, which diffs this against a
real Device_Tampering_Risk_Report_v2.xlsx.
"""

import csv
import math
from datetime import datetime
from collections import defaultdict

# ----------------------------------------------------------------------
# Tunable settings - identical to tampering_report.py
# ----------------------------------------------------------------------
GAP_DISTANCE_THRESHOLD_KM = 2.0
MAX_PLAUSIBLE_SPEED_KMH = 140.0
NULL_FIX_TOLERANCE_DEG = 0.001
DATE_TIME_FORMAT = "%d/%m/%Y %H:%M:%S"

REQUIRED_MOVEMENT_COLS = [
    "AssetID", "AssetExtra", "FleetNumber", "DepartureDate", "DepartureTime",
    "ArrivalDate", "ArrivalTime", "StartLatLong", "EndLatLong",
    "DepartFrom", "ArriveAt", "StartOdoMeter", "EndOdoMeter",
]
REQUIRED_EVENT_COLS = ["AssetID", "EventDescription", "EventStartDate", "EventStartTime"]


def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def parse_latlong(raw):
    if not raw or not raw.strip():
        return None
    s = raw.strip()
    parts = [p.strip() for p in s.replace(",", "/").split("/")]
    if len(parts) < 2:
        return None
    try:
        lat, lon = float(parts[0]), float(parts[1])
    except ValueError:
        return None
    if abs(lat) < NULL_FIX_TOLERANCE_DEG and abs(lon) < NULL_FIX_TOLERANCE_DEG:
        return None
    return lat, lon


def parse_dt(date_str, time_str):
    return datetime.strptime(date_str.strip() + " " + time_str.strip(), DATE_TIME_FORMAT)


def parse_odo(raw):
    if not raw or not raw.strip():
        return None
    try:
        return float(raw.strip().replace(",", ""))
    except ValueError:
        return None


def clean_fleet_number(raw):
    return (raw or "").strip().strip(",")


def readable_duration(total_minutes):
    if total_minutes is None:
        return ""
    total_min = int(round(total_minutes))
    days, rem = divmod(total_min, 1440)
    hrs, mins = divmod(rem, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hrs:
        parts.append(f"{hrs}h")
    parts.append(f"{mins}m")
    return " ".join(parts)


def severity(distance_km):
    if distance_km > 100:
        return "Critical"
    if distance_km >= 10:
        return "High"
    return "Moderate"


def read_csv_rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def validate_columns(rows, required, label, path):
    if not rows:
        raise ValueError(f"{label} file is empty: {path}")
    missing = [c for c in required if c not in rows[0]]
    if missing:
        raise ValueError(
            f"{label} file '{path}' is missing expected column(s): {', '.join(missing)}\n"
            f"Columns found: {', '.join(rows[0].keys())}"
        )


def analyse(movement_csv_path, event_csv_path):
    """Identical logic and return shape to tampering_report.py's analyse()."""
    trips = read_csv_rows(movement_csv_path)
    validate_columns(trips, REQUIRED_MOVEMENT_COLS, "Daily Movement Report", movement_csv_path)

    events = read_csv_rows(event_csv_path)
    validate_columns(events, REQUIRED_EVENT_COLS, "Detailed Event Report", event_csv_path)

    by_asset = defaultdict(list)
    bad_trip_rows = 0
    for t in trips:
        try:
            dep_dt = parse_dt(t["DepartureDate"], t["DepartureTime"])
            arr_dt = parse_dt(t["ArrivalDate"], t["ArrivalTime"])
        except Exception:
            bad_trip_rows += 1
            continue
        by_asset[t["AssetID"]].append((dep_dt, arr_dt, t))

    events_by_asset = defaultdict(list)
    bad_event_rows = 0
    for e in events:
        try:
            edt = parse_dt(e["EventStartDate"], e["EventStartTime"])
        except Exception:
            bad_event_rows += 1
            continue
        events_by_asset[e["AssetID"]].append((edt, e))
    for a in events_by_asset:
        events_by_asset[a].sort(key=lambda x: x[0])

    def power_events_in_window(asset_id, start_dt, end_dt):
        return [(edt, e["EventDescription"]) for edt, e in events_by_asset.get(asset_id, [])
                if start_dt <= edt <= end_dt]

    gaps = []
    skipped = []

    for aid, tlist in by_asset.items():
        tlist.sort(key=lambda x: x[0])
        fleet = clean_fleet_number(tlist[0][2]["FleetNumber"])

        for i in range(len(tlist) - 1):
            dep_i, arr_i, trip_i = tlist[i]
            dep_j, arr_j, trip_j = tlist[i + 1]

            loc_end = parse_latlong(trip_i["EndLatLong"])
            loc_start = parse_latlong(trip_j["StartLatLong"])

            if loc_end is None or loc_start is None:
                skipped.append({
                    "AssetID": aid,
                    "AssetName": trip_i["AssetExtra"].strip(),
                    "FleetNumber": fleet,
                    "ArrivalDate": trip_i["ArrivalDate"],
                    "ArrivalTime": trip_i["ArrivalTime"],
                    "RawEndLatLong": trip_i["EndLatLong"],
                    "NextDepartureDate": trip_j["DepartureDate"],
                    "NextDepartureTime": trip_j["DepartureTime"],
                    "RawStartLatLong": trip_j["StartLatLong"],
                })
                continue

            dist = haversine(loc_end[0], loc_end[1], loc_start[0], loc_start[1])
            gap_minutes = (dep_j - arr_i).total_seconds() / 60.0
            if gap_minutes < 0:
                continue

            odo_end_i = parse_odo(trip_i["EndOdoMeter"])
            odo_start_j = parse_odo(trip_j["StartOdoMeter"])
            odo_jump = (odo_start_j - odo_end_i) if (odo_end_i is not None and odo_start_j is not None) else None

            implied_speed = dist / (gap_minutes / 60.0) if gap_minutes > 0 else None

            hits = power_events_in_window(aid, arr_i, dep_j)
            has_disconnect = any(desc == "Power Disconnect" for _, desc in hits)
            has_reconnect = any(desc == "Power Reconnect" for _, desc in hits)

            gaps.append({
                "AssetID": aid,
                "AssetName": trip_i["AssetExtra"].strip(),
                "FleetNumber": fleet,
                "ArrivalDate": trip_i["ArrivalDate"],
                "ArrivalTime": trip_i["ArrivalTime"],
                "ArriveAt": trip_i["ArriveAt"],
                "NextDepartureDate": trip_j["DepartureDate"],
                "NextDepartureTime": trip_j["DepartureTime"],
                "DepartFrom": trip_j["DepartFrom"],
                "DistanceKm": round(dist, 2),
                "GapMinutes": round(gap_minutes, 1),
                "OdoJumpKm": round(odo_jump, 2) if odo_jump is not None else None,
                "ImpliedSpeedKmh": round(implied_speed, 1) if implied_speed is not None else None,
                "HasDisconnect": has_disconnect,
                "HasReconnect": has_reconnect,
                "PowerEventCount": len(hits),
                "PlausibleSpeed": (implied_speed is None) or (implied_speed <= MAX_PLAUSIBLE_SPEED_KMH),
            })

    mismatches = [g for g in gaps if g["DistanceKm"] > GAP_DISTANCE_THRESHOLD_KM]
    confirmed_power_cycle = [g for g in mismatches if g["HasDisconnect"] and g["HasReconnect"]]
    confirmed_no_reconnect = [g for g in mismatches if g["HasDisconnect"] and not g["HasReconnect"]]
    confirmed = confirmed_power_cycle + confirmed_no_reconnect
    unconfirmed = [g for g in mismatches if not g["HasDisconnect"]]

    return {
        "total_trip_records": len(trips),
        "total_assets": len(by_asset),
        "bad_trip_rows": bad_trip_rows,
        "bad_event_rows": bad_event_rows,
        "gaps": gaps,
        "mismatches": mismatches,
        "confirmed": confirmed,
        "confirmed_power_cycle": confirmed_power_cycle,
        "confirmed_no_reconnect": confirmed_no_reconnect,
        "unconfirmed": unconfirmed,
        "skipped": skipped,
    }


# ----------------------------------------------------------------------
# Workbook building - ported verbatim from the original standalone tool
# (tampering_report.py's build_workbook()/write_case_sheet()), same
# styling, same sheet layout, same everything. This is what makes the
# "Tampering Risk Report" export button on the dashboard produce a real
# file instead of a 404.
# ----------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT_NAME = "Arial"
NAVY = "1F3864"
BLUE = "2E5395"
WHITE = "FFFFFF"
GREY = "808080"
RED = "C00000"
RED_FILL = "F8CBAD"
ORANGE_FILL = "FCE4D6"
YELLOW_FILL = "FFF2CC"
SEV_FILL = {"Critical": RED_FILL, "High": ORANGE_FILL, "Moderate": YELLOW_FILL}
SEV_FONT = {"Critical": RED, "High": "C65911", "Moderate": "7F6000"}


def style_header(ws, row, ncols, fill=NAVY):
    ws.row_dimensions[row].height = 24
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def title_block(ws, title, subtitle, ncols, start_row=1):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    c = ws.cell(row=start_row, column=1, value=title)
    c.font = Font(name=FONT_NAME, bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row].height = 30

    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=ncols)
    c2 = ws.cell(row=start_row + 1, column=1, value=subtitle)
    c2.font = Font(name=FONT_NAME, italic=True, size=10, color=WHITE)
    c2.fill = PatternFill("solid", fgColor=BLUE)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row + 1].height = 18


def write_case_sheet(wb, name, records, subtitle):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    headers = ["Asset ID", "Vehicle", "Fleet Number", "Severity", "Arrival Date", "Arrival Time",
               "Last Known Location", "Next Departure Date", "Next Departure Time", "Next Departure Location",
               "Gap Distance (km)", "Gap Duration", "Implied Speed (km/h)", "Power Event In Gap"]
    ncols = len(headers)
    title_block(ws, name, subtitle, ncols)
    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=1 + i, value=h)
    style_header(ws, hr, ncols)

    r = hr
    for g in sorted(records, key=lambda x: -x["DistanceKm"]):
        r += 1
        s = severity(g["DistanceKm"])
        if g["HasDisconnect"] and g["HasReconnect"]:
            pe = "Disconnect + Reconnect logged (Power Cycle)"
        elif g["HasDisconnect"]:
            pe = "Disconnect logged, no Reconnect"
        elif g["HasReconnect"]:
            pe = "Reconnect only, no Disconnect"
        else:
            pe = "None logged"
        vals = [g["AssetID"], g["AssetName"], g["FleetNumber"], s, g["ArrivalDate"], g["ArrivalTime"],
                g["ArriveAt"], g["NextDepartureDate"], g["NextDepartureTime"], g["DepartFrom"],
                g["DistanceKm"], readable_duration(g["GapMinutes"]), g["ImpliedSpeedKmh"], pe]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=1 + i, value=v).font = Font(name=FONT_NAME, size=9.5)
        fill = PatternFill("solid", fgColor=SEV_FILL[s])
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=4).font = Font(name=FONT_NAME, size=9.5, bold=True, color=SEV_FONT[s])
        ws.cell(row=r, column=11).font = Font(name=FONT_NAME, size=9.5, bold=True)

    widths = [10, 26, 20, 10, 13, 11, 24, 15, 13, 24, 15, 14, 15, 30]
    for i, wd in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = wd
    ws.freeze_panes = f"A{hr + 1}"

    if r > hr:
        safe_name = "".join(ch for ch in name if ch.isalnum())
        tab = Table(displayName=safe_name, ref=f"A{hr}:{get_column_letter(ncols)}{r}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
        ws.add_table(tab)
    return r - hr


def build_workbook(results, output_path):
    """Same workbook, same styling, same sheet layout as the original
    standalone Device Tampering Risk Report tool. Identical output
    whether this ran from the GUI tool or from the daily import."""
    gaps = results["gaps"]
    mismatches = results["mismatches"]
    confirmed = results["confirmed"]
    confirmed_power_cycle = results["confirmed_power_cycle"]
    confirmed_no_reconnect = results["confirmed_no_reconnect"]
    unconfirmed = results["unconfirmed"]
    skipped = results["skipped"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ncols = 5
    title_block(ws, "Device Tampering Risk Report - Location Continuity Check",
                "Method: arrival location of each trip vs. departure location of that vehicle's next trip.", ncols)

    row = 4
    labels = ["Consecutive Trip Gaps Checked", "Location Mismatches (>2km)",
              "Confirmed (Power Event Logged)", "Unconfirmed (No Power Event Logged)"]
    values = [str(len(gaps)), str(len(mismatches)), str(len(confirmed)), str(len(unconfirmed))]
    colors = [BLUE, "C65911", RED, "7F6000"]
    for i in range(4):
        lab = ws.cell(row=row, column=1 + i, value=labels[i])
        lab.font = Font(name=FONT_NAME, size=8.5, bold=True, color=WHITE)
        lab.fill = PatternFill("solid", fgColor=colors[i])
        lab.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        val = ws.cell(row=row + 1, column=1 + i, value=values[i])
        val.font = Font(name=FONT_NAME, size=16, bold=True, color=colors[i])
        val.fill = PatternFill("solid", fgColor="F2F2F2")
        val.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34
    ws.row_dimensions[row + 1].height = 30

    row += 3
    ws.cell(row=row, column=1, value="Method").font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
    row += 1
    method_lines = [
        "1. Every trip's arrival location is compared to the departure location of that vehicle's next trip.",
        "2. If the two points are more than 2 km apart, the vehicle relocated without a trip being logged. Primary flag.",
        "3. Each flagged gap is checked against Power Disconnect / Power Reconnect events for that vehicle in the same window.",
        "   - CONFIRMED (Power Cycle): both a Disconnect and a Reconnect fall inside the gap. Strongest evidence.",
        "   - CONFIRMED (Disconnect, No Reconnect): a Disconnect was logged but never reconnected. Device may still be offline.",
        "   - UNCONFIRMED: no power event logged at all. Vehicle still moved untracked - needs a field check",
        "     (device removed/damaged, SIM/antenna fault, or a signal blackout are also possible causes).",
        f"4. Implied average speed is sanity-checked against {MAX_PLAUSIBLE_SPEED_KMH:.0f} km/h; gaps above that are still shown",
        "   but should be reviewed for a bad GPS fix before being treated as confirmed movement.",
    ]
    for line in method_lines:
        ws.cell(row=row, column=1, value=line).font = Font(name=FONT_NAME, size=9.5)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Data Quality - Excluded From Analysis").font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
    row += 1
    dq_lines = [
        f"{len(skipped)} trip boundaries were excluded because one side had a null GPS fix (0,0) - a device placeholder",
        "for 'no signal', not a real coordinate. See the 'Data Quality Log' tab for the full list.",
    ]
    if results["bad_trip_rows"] or results["bad_event_rows"]:
        dq_lines.append(
            f"{results['bad_trip_rows']} movement rows and {results['bad_event_rows']} event rows had unreadable "
            f"dates/times and were skipped entirely."
        )
    for line in dq_lines:
        ws.cell(row=row, column=1, value=line).font = Font(name=FONT_NAME, size=9.5, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Severity Bands").font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
    row += 1
    headers = ["Severity", "Gap Distance", "Confirmed", "Unconfirmed"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=1 + i, value=h)
    style_header(ws, row, 4, fill=BLUE)
    for sname, rule in [("Critical", "> 100 km"), ("High", "10 - 100 km"), ("Moderate", "2 - 10 km")]:
        row += 1
        cconf = sum(1 for g in confirmed if severity(g["DistanceKm"]) == sname)
        cunconf = sum(1 for g in unconfirmed if severity(g["DistanceKm"]) == sname)
        ws.cell(row=row, column=1, value=sname).font = Font(name=FONT_NAME, bold=True, color=SEV_FONT[sname])
        ws.cell(row=row, column=2, value=rule).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=row, column=3, value=cconf).font = Font(name=FONT_NAME, size=10, bold=True)
        ws.cell(row=row, column=4, value=cunconf).font = Font(name=FONT_NAME, size=10, bold=True)
        fill = PatternFill("solid", fgColor=SEV_FILL[sname])
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = fill

    row += 2
    ws.cell(row=row, column=1, value="Confirmed Case Breakdown").font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
    row += 1
    ws.cell(row=row, column=1, value="Power Cycle (Disconnect + Reconnect)").font = Font(name=FONT_NAME, size=10)
    ws.cell(row=row, column=2, value=len(confirmed_power_cycle)).font = Font(name=FONT_NAME, size=10, bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Disconnect, No Reconnect (device may still be offline)").font = Font(name=FONT_NAME, size=10)
    ws.cell(row=row, column=2, value=len(confirmed_no_reconnect)).font = Font(name=FONT_NAME, size=10, bold=True)

    row += 2
    ws.cell(row=row, column=1, value="Top Vehicles - Confirmed Cases").font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
    row += 1
    headers2 = ["Asset ID", "Vehicle", "Fleet Number", "Confirmed Cases", "Worst Gap (km)"]
    for i, h in enumerate(headers2):
        ws.cell(row=row, column=1 + i, value=h)
    style_header(ws, row, 5, fill=RED)
    vstat = {}
    for g in confirmed:
        k = g["AssetID"]
        vstat.setdefault(k, {"name": g["AssetName"], "fleet": g["FleetNumber"], "n": 0, "max": 0})
        vstat[k]["n"] += 1
        vstat[k]["max"] = max(vstat[k]["max"], g["DistanceKm"])
    for k, v in sorted(vstat.items(), key=lambda x: -x[1]["max"]):
        row += 1
        ws.cell(row=row, column=1, value=k).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=row, column=2, value=v["name"]).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=row, column=3, value=v["fleet"]).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=row, column=4, value=v["n"]).font = Font(name=FONT_NAME, size=10, bold=True)
        ws.cell(row=row, column=5, value=round(v["max"], 1)).font = Font(name=FONT_NAME, size=10, bold=True, color=RED)

    for i, wdt in enumerate([18, 30, 22, 20, 55]):
        ws.column_dimensions[get_column_letter(i + 1)].width = wdt

    write_case_sheet(
        wb, "Confirmed Cases", confirmed,
        "Location gap AND a Power Disconnect event logged in the same window. Direct evidence of tampering during an untracked move."
    )
    write_case_sheet(
        wb, "Unconfirmed Cases", unconfirmed,
        "Location gap with NO Power Disconnect logged. Vehicle still moved untracked - needs a field check on device, SIM and antenna."
    )

    wsq = wb.create_sheet("Data Quality Log")
    wsq.sheet_view.showGridLines = False
    headers = ["Asset ID", "Vehicle", "Fleet Number", "Arrival Date", "Arrival Time", "Raw End Coordinate",
               "Next Departure Date", "Next Departure Time", "Raw Start Coordinate"]
    title_block(wsq, "Data Quality Log - Null GPS Fixes Excluded",
                f"{len(skipped)} trip boundaries excluded because one side reported a null (0,0) coordinate.", len(headers))
    hr = 4
    for i, h in enumerate(headers):
        wsq.cell(row=hr, column=1 + i, value=h)
    style_header(wsq, hr, len(headers), fill=GREY)
    r = hr
    for s in sorted(skipped, key=lambda x: (x["AssetID"], x["ArrivalDate"])):
        r += 1
        vals = [s["AssetID"], s["AssetName"], s["FleetNumber"], s["ArrivalDate"], s["ArrivalTime"],
                s["RawEndLatLong"], s["NextDepartureDate"], s["NextDepartureTime"], s["RawStartLatLong"]]
        for i, v in enumerate(vals):
            wsq.cell(row=r, column=1 + i, value=v).font = Font(name=FONT_NAME, size=9)
    widths = [10, 26, 20, 13, 11, 20, 15, 13, 20]
    for i, wd in enumerate(widths):
        wsq.column_dimensions[get_column_letter(i + 1)].width = wd
    wsq.freeze_panes = f"A{hr + 1}"

    wb.save(output_path)
    return output_path
