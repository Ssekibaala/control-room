"""
New sheet structure, per direct client feedback:

  1. Executive Dashboard   - KPIs (formula-driven), settings snapshot, headline insights
  2. Critical Assets        - ranked worst-first, the action list
  3. Pending Customer Feedback - awaiting confirmation, with feedback status if any
  4. Border Risk             - only genuine offline+near-border cases
  5. Recovered Since Yesterday - day-over-day comparison (empty with a note on first run)
  6. Healthy Fleet            - separated out, reference only, not mixed with issues
  7. Settings                 - what thresholds produced this exact report
  8. Full Data                - every asset, audit trail, autofilter

No confidence score anywhere. Each platform's status is shown
individually. No Location Mismatch column, the per-platform table and
detail/reasons columns already carry that information.
"""

from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT_NAME = "Arial"
NAVY = "1F3864"
STEEL = "2E5C8A"
WHITE = "FFFFFF"
CRITICAL_RED = "C00000"
HIGH_ORANGE = "ED7D31"
ELEVATED_YELLOW = "FFC000"
ONLINE_FILL = "E2EFDA"
CRITICAL_FILL = "FFC7CE"
HIGH_FILL = "FCE4D6"
ELEVATED_FILL = "FFF2CC"
PENDING_FILL = "DDEBF7"
BORDER_FILL = "F4CCCC"
GREY_FILL = "F2F2F2"

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALL_PLATFORMS = ["Teletrac", "MiX Unity", "FT Cloud Camera"]

FULL_HEADERS = [
    "Asset Plate", "Status", "Severity", "Days Offline",
    "Teletrac", "MiX Unity", "FT Cloud Camera",
    "Border Risk", "Nearest Border", "Distance (km)",
    "Customer Feedback", "Recommended Action",
    "Investigation Reasons", "Last Known Location",
]


def _style_header(ws, row_idx, ncols, fill=NAVY):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(name=FONT_NAME, color=WHITE, bold=True, size=11)
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX


def _severity_fill(severity, status):
    if status == "Online":
        return ONLINE_FILL
    if severity == "Critical - Long-term Fault":
        return CRITICAL_FILL
    if severity == "High - Escalate This Week":
        return HIGH_FILL
    if severity == "Elevated - Monitor":
        return ELEVATED_FILL
    return PENDING_FILL


def _row_for(plate, info):
    ps = info["platform_status"]
    platform_cells = []
    for p in ALL_PLATFORMS:
        if p in ps:
            platform_cells.append(ps[p][0])  # "Online" / "Offline" / "No Data"
        else:
            platform_cells.append("N/A")

    border_name, border_country, border_dist = (None, None, None)
    if info["border_detail"]:
        border_name, border_country, border_dist = info["border_detail"]

    fb = info["feedback"]
    fb_text = f"{fb['status']}: {fb['comment']}" if fb else ""

    return [
        plate, info["status"], info["severity"], info["days_silent"],
        *platform_cells,
        "Yes" if info["border_flag"] else "No",
        f"{border_name} ({border_country})" if border_name else "",
        border_dist if border_dist is not None else "",
        fb_text,
        info["action"],
        "; ".join(info["reasons"]),
        info.get("last_location") or "",
    ]


def _write_row(ws, plate, info):
    ws.append(_row_for(plate, info))
    r = ws.max_row
    for c in range(1, len(FULL_HEADERS) + 1):
        ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=r, column=c).border = BOX
        ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=False)
    fill = _severity_fill(info["severity"], info["status"])
    ws.cell(row=r, column=2).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    if info["border_flag"]:
        c = ws.cell(row=r, column=8)
        c.fill = PatternFill(start_color=BORDER_FILL, end_color=BORDER_FILL, fill_type="solid")
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=CRITICAL_RED)
    return r


_COL_WIDTHS = {'A': 14, 'B': 24, 'C': 24, 'D': 12, 'E': 12, 'F': 12, 'G': 14,
               'H': 12, 'I': 20, 'J': 12, 'K': 34, 'L': 32, 'M': 46, 'N': 36}


def _autosize(ws):
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = w


def write_report(results: dict, settings: dict, recovered: list, newly_offline: list,
                  output_path: str, report_date=None, history_available=True):
    report_date = report_date or datetime.now()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    online_plates = [p for p, v in results.items() if v["status"] == "Online"]
    critical_plates = sorted(
        [p for p, v in results.items() if v["status"] == "Technical Escalation"],
        key=lambda p: -results[p]["days_silent"]
    )
    pending_plates = sorted(
        [p for p, v in results.items() if v["status"] == "Pending Customer Confirmation"],
        key=lambda p: -results[p]["days_silent"]
    )
    border_plates = [p for p in critical_plates + pending_plates if results[p]["border_flag"]]

    # ================= FULL DATA (built first, dashboard formulas reference it) =================
    data_ws = wb.create_sheet("Full Data")
    data_ws.append(FULL_HEADERS)
    _style_header(data_ws, 1, len(FULL_HEADERS))
    data_ws.freeze_panes = "A2"
    for plate in sorted(results.keys(), key=lambda p: -results[p]["days_silent"]):
        _write_row(data_ws, plate, results[plate])
    last_row = data_ws.max_row
    _autosize(data_ws)
    table = Table(displayName="FleetData", ref=f"A1:{get_column_letter(len(FULL_HEADERS))}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    data_ws.add_table(table)

    # ================= EXECUTIVE DASHBOARD =================
    ws = wb.create_sheet("Executive Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "GTL FLEET INTEGRITY REPORT"
    ws["A1"].font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Automated Cross-Platform Analysis  |  Generated {report_date.strftime('%d %B %Y, %H:%M')}"
    ws["A2"].font = Font(name=FONT_NAME, size=11, italic=True, color="595959")
    ws.merge_cells("A3:F3")
    ws["A3"] = "Sources: Teletrac, MiX Unity (Mobile Status, Movement, Power Events), FT Cloud Camera Platform"
    ws["A3"].font = Font(name=FONT_NAME, size=9, italic=True, color="808080")

    row = 5
    ws.cell(row=row, column=1, value="FLEET HEALTH OVERVIEW").font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    kpis = [
        ("Total Assets Tracked", f"=COUNTA('Full Data'!A2:A{last_row})"),
        ("Online / Healthy", f"=COUNTIF('Full Data'!B2:B{last_row},\"Online\")"),
        ("Pending Customer Confirmation", f"=COUNTIF('Full Data'!B2:B{last_row},\"Pending Customer Confirmation\")"),
        ("Technical Escalation (Total)", f"=COUNTIF('Full Data'!B2:B{last_row},\"Technical Escalation\")"),
        ("  Critical - Long-term Fault", f"=COUNTIF('Full Data'!C2:C{last_row},\"Critical - Long-term Fault\")"),
        ("  High - Escalate This Week", f"=COUNTIF('Full Data'!C2:C{last_row},\"High - Escalate This Week\")"),
        ("  Elevated - Monitor", f"=COUNTIF('Full Data'!C2:C{last_row},\"Elevated - Monitor\")"),
        ("Genuine Border-Crossing Risk", f"=COUNTIF('Full Data'!H2:H{last_row},\"Yes\")"),
    ]
    kpi_first_row = row
    for label, formula in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT_NAME, size=10.5, bold=not label.startswith(" "))
        ws.cell(row=row, column=2, value=formula).font = Font(name=FONT_NAME, size=10.5, bold=True, color=NAVY)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    health_row = row + 1
    ws.cell(row=health_row, column=1, value="Fleet Health Rate").font = Font(name=FONT_NAME, size=12, bold=True)
    ws.cell(row=health_row, column=2, value=f"=B{kpi_first_row+1}/B{kpi_first_row}").font = Font(
        name=FONT_NAME, size=12, bold=True, color=NAVY)
    ws.cell(row=health_row, column=2).number_format = "0.0%"
    ws.cell(row=health_row, column=2).alignment = Alignment(horizontal="center")
    row = health_row + 2

    # Recovery / new-offline callout
    ws.cell(row=row, column=1, value="DAY-OVER-DAY CHANGE").font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).fill = PatternFill(start_color=STEEL, end_color=STEEL, fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1
    if history_available:
        ws.cell(row=row, column=1, value=f"Recovered since yesterday: {len(recovered)}").font = Font(name=FONT_NAME, size=10.5)
        row += 1
        ws.cell(row=row, column=1, value=f"Newly offline since yesterday: {len(newly_offline)}").font = Font(name=FONT_NAME, size=10.5)
        row += 1
    else:
        ws.cell(row=row, column=1,
                value="No prior-day snapshot yet, day-over-day tracking starts from tomorrow's run.").font = Font(
            name=FONT_NAME, size=10.5, italic=True, color="808080")
        row += 1
    row += 1

    # Callout + insights
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="TOP PRIORITY: see 'Critical Assets' tab for the ranked action list, worst case first.")
    ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill(start_color=CRITICAL_RED, end_color=CRITICAL_RED, fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 2

    critical_count = sum(1 for v in results.values() if v.get("severity") == "Critical - Long-term Fault")
    insights = [
        f"{len(border_plates)} asset(s) show genuine border-crossing risk: offline AND last seen within "
        f"{settings['BORDER_RADIUS_KM']:.0f}km of a known crossing. See 'Border Risk' tab.",
        f"{critical_count} asset(s) have been offline {settings['LONG_TERM_FAULT_DAYS']}+ days on every platform, "
        "these are the longest-standing faults, recommend field recovery.",
        f"Offline threshold in effect for this report: {settings['OFFLINE_THRESHOLD_DAYS']} day(s), "
        f"editable in {settings['_path']}.",
    ]
    for line in insights:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"\u2022  {line}")
        cell.font = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    _autosize_dashboard(ws)

    # ================= CRITICAL ASSETS =================
    crit_ws = wb.create_sheet("Critical Assets")
    crit_ws.append(["Rank"] + FULL_HEADERS)
    _style_header(crit_ws, 1, len(FULL_HEADERS) + 1, fill=CRITICAL_RED)
    crit_ws.freeze_panes = "A2"
    for rank, plate in enumerate(critical_plates, start=1):
        crit_ws.append([rank] + _row_for(plate, results[plate]))
        r = crit_ws.max_row
        for c in range(1, len(FULL_HEADERS) + 2):
            crit_ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
            crit_ws.cell(row=r, column=c).border = BOX
        fill = _severity_fill(results[plate]["severity"], results[plate]["status"])
        crit_ws.cell(row=r, column=3).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    _autosize_shifted(crit_ws)

    # ================= PENDING CUSTOMER FEEDBACK =================
    pend_ws = wb.create_sheet("Pending Customer Feedback")
    pend_ws.append(FULL_HEADERS)
    _style_header(pend_ws, 1, len(FULL_HEADERS), fill=STEEL)
    pend_ws.freeze_panes = "A2"
    for plate in pending_plates:
        _write_row(pend_ws, plate, results[plate])
    _autosize(pend_ws)

    # ================= BORDER RISK =================
    border_ws = wb.create_sheet("Border Risk")
    border_ws.append(FULL_HEADERS)
    _style_header(border_ws, 1, len(FULL_HEADERS), fill=CRITICAL_RED)
    border_ws.freeze_panes = "A2"
    for plate in sorted(border_plates, key=lambda p: results[p]["border_detail"][2]):
        _write_row(border_ws, plate, results[plate])
    if not border_plates:
        border_ws.append(["No genuine border-crossing risk cases in this run."] + [""] * (len(FULL_HEADERS) - 1))
        border_ws.cell(row=2, column=1).font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
    _autosize(border_ws)

    # ================= RECOVERED SINCE YESTERDAY =================
    rec_ws = wb.create_sheet("Recovered Since Yesterday")
    rec_ws.append(["Asset Plate", "Current Status"])
    _style_header(rec_ws, 1, 2, fill=STEEL)
    if not history_available:
        rec_ws.append(["No prior-day snapshot yet.", "Starts tracking from tomorrow's run."])
        rec_ws.cell(row=2, column=1).font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
    else:
        for plate in recovered:
            rec_ws.append([plate, results.get(plate, {}).get("status", "Online")])
        if not recovered:
            rec_ws.append(["No assets recovered since yesterday.", ""])
            rec_ws.cell(row=2, column=1).font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
    rec_ws.column_dimensions['A'].width = 18
    rec_ws.column_dimensions['B'].width = 26

    # ================= HEALTHY FLEET =================
    healthy_ws = wb.create_sheet("Healthy Fleet")
    healthy_ws.append(["Asset Plate", "Teletrac", "MiX Unity", "FT Cloud Camera", "Last Known Location"])
    _style_header(healthy_ws, 1, 5, fill="548235")
    healthy_ws.freeze_panes = "A2"
    for plate in sorted(online_plates):
        info = results[plate]
        ps = info["platform_status"]
        row_vals = [plate] + [ps.get(p, ("N/A", None))[0] for p in ALL_PLATFORMS] + [info.get("last_location") or ""]
        healthy_ws.append(row_vals)
        r = healthy_ws.max_row
        for c in range(1, 6):
            healthy_ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
            healthy_ws.cell(row=r, column=c).border = BOX
    for col, w in zip("ABCDE", [14, 12, 12, 14, 42]):
        healthy_ws.column_dimensions[col].width = w

    # ================= SETTINGS =================
    set_ws = wb.create_sheet("Settings")
    set_ws.append(["Setting", "Value Used For This Report"])
    _style_header(set_ws, 1, 2, fill=NAVY)
    for label, key in [
        ("Offline Threshold (days)", "OFFLINE_THRESHOLD_DAYS"),
        ("High Priority Threshold (days)", "HIGH_PRIORITY_DAYS"),
        ("Long-term Fault Threshold (days)", "LONG_TERM_FAULT_DAYS"),
        ("Border Radius (km)", "BORDER_RADIUS_KM"),
        ("Ignore Demo Vehicles", "IGNORE_DEMO_VEHICLES"),
    ]:
        set_ws.append([label, settings[key]])
    set_ws.append([])
    set_ws.append(["Settings file location", settings["_path"]])
    set_ws.append(["To change these values", "Edit settings.ini next to the EXE, then re-run. No code changes needed."])
    set_ws.column_dimensions['A'].width = 34
    set_ws.column_dimensions['B'].width = 60

    wb.save(output_path)
    return output_path


def _autosize_dashboard(ws):
    for col, w in {'A': 40, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16}.items():
        ws.column_dimensions[col].width = w


def _autosize_shifted(ws):
    widths = {'A': 6}
    letters = list(_COL_WIDTHS.keys())
    for i, letter in enumerate(letters):
        shifted = get_column_letter(i + 2)
        widths[shifted] = _COL_WIDTHS[letter]
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_email_body(results: dict, settings: dict, recovered, newly_offline, report_date=None,
                      history_available=True):
    report_date = report_date or datetime.now()
    total = len(results)
    online = sum(1 for v in results.values() if v["status"] == "Online")
    pending = [p for p, v in results.items() if v["status"] == "Pending Customer Confirmation"]
    critical = sorted(
        [p for p, v in results.items() if v["status"] == "Technical Escalation"],
        key=lambda p: -results[p]["days_silent"]
    )
    border = [p for p in critical + pending if results[p]["border_flag"]]
    health_rate = f"{round(online / total * 100)}%" if total else "0%"

    lines = []
    lines.append("Dear Justin/Brian,")
    lines.append("")
    lines.append(
        "Please find below the daily GTL fleet integrity status update, automated cross-platform "
        f"analysis across MiX Unity, Teletrac, and FT Cloud, generated {report_date.strftime('%d %B %Y %H:%M')}."
    )
    lines.append("")
    lines.append(f"Total Assets Tracked: {total}")
    lines.append(f"Online / Healthy: {online}")
    lines.append(f"Pending Customer Confirmation: {len(pending)}")
    lines.append(f"Technical Escalation: {len(critical)}")
    lines.append(f"Fleet Health Rate: {health_rate}")
    lines.append(f"Genuine Border-Crossing Risk: {len(border)}")
    if history_available:
        lines.append(f"Recovered Since Yesterday: {len(recovered)}   |   Newly Offline: {len(newly_offline)}")
    lines.append("")
    if critical:
        lines.append("Top Priority (worst-first, full ranked list with recommended actions in workbook):")
        for p in critical[:10]:
            info = results[p]
            lines.append(f"  {info['days_silent']:>6}d  {p}  [{info['severity']}]  action: {info['action']}")
        if len(critical) > 10:
            lines.append(f"  ...and {len(critical) - 10} more, see Critical Assets tab.")
        lines.append("")
    lines.append(f"Thresholds in effect: offline after {settings['OFFLINE_THRESHOLD_DAYS']} day(s), "
                  f"long-term fault after {settings['LONG_TERM_FAULT_DAYS']} day(s). Editable in settings.ini.")
    lines.append("")
    lines.append("Regards,")
    lines.append("Teletrac Fleets Solutions - Technical Team")
    return "\n".join(lines)
