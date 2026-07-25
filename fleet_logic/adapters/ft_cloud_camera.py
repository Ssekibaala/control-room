"""
Adapter for: Offline Vehicle Reports (FT Cloud camera platform)
Arrives as an xlsx inside a zip attachment.

Confirmed real columns: License Plate Number, Fleets, Drivers,
Last Online Time, Offline Time, Longitude, Latitude, Last Positioning

Only rows where Fleets contains 'Globe Trotters' or '(GTL)' belong to
this client, same filtering pattern as every other source.
"""

import zipfile
import io
import openpyxl
from datetime import datetime
from schema import AssetReport, normalize_plate

GTL_FLEET_MARKERS = ("Globe Trotters", "GTL")


def _is_gtl(fleet_name: str) -> bool:
    return any(marker.lower() in (fleet_name or "").lower() for marker in GTL_FLEET_MARKERS)


def _parse_timestamp(ts_str):
    if not ts_str:
        return None
    if isinstance(ts_str, datetime):
        return ts_str
    try:
        return datetime.strptime(str(ts_str).strip(), "%d/%m/%Y %H:%M:%S")
    except ValueError:
        return None


def parse(zip_path: str):
    rows = []
    with zipfile.ZipFile(zip_path) as z:
        xlsx_name = next(n for n in z.namelist() if n.endswith('.xlsx'))
        with z.open(xlsx_name) as f:
            data = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(data, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        fleet = row[idx.get('Fleets', -1)] if idx.get('Fleets') is not None else None
        if not _is_gtl(fleet):
            continue
        plate_raw = row[idx.get('License Plate Number', -1)]
        plate = normalize_plate(plate_raw)
        if not plate:
            continue
        lon = row[idx.get('Longitude', -1)] if 'Longitude' in idx else None
        lat = row[idx.get('Latitude', -1)] if 'Latitude' in idx else None
        rows.append(AssetReport(
            asset_plate=plate,
            source_platform="FT Cloud Camera",
            report_type="camera_status",
            last_report_time=_parse_timestamp(row[idx.get('Last Online Time', -1)]),
            last_lat=float(lat) if lat not in (None, "") else None,
            last_lon=float(lon) if lon not in (None, "") else None,
            last_location_text=row[idx.get('Last Positioning', -1)] if 'Last Positioning' in idx else None,
            status_note=row[idx.get('Offline Time', -1)] if 'Offline Time' in idx else None,
            raw_row={h: row[i] for h, i in idx.items()},
        ))
    return rows
