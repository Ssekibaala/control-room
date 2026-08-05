"""
Admin-only billing helper: every MiX asset for one or more chosen
organisations, with its install date and who installed it, so assets
added to the platform in a given month can be identified and billed.
Fetched on demand from the admin panel, not on the regular poll cadence
- this is an occasional lookup, not a live dashboard figure.

Install date = MiX's "CreatedDate" field, installer = MiX's "CreatedBy"
field, both on the asset record itself - confirmed live against real
data (orgs AGL, Hardware World, SRZ, VENGAR ENERGY, 314 assets total,
all four platform clients checked): CreatedDate is the ONLY date-shaped
field MiX's /assets/group/{orgId} response ever returns, and both
fields were present on every asset checked, none missing.

No IMEI/device-type field exists in this account's MiX data: checked
the full Integrate API surface (185 endpoints via /swagger/docs/v1),
the Asset schema's SerialNumber/AdditionalMobileDevice fields (null on
every asset sampled), and the per-org custom-fields endpoint
(/assets/group/{id}/additionaldetails, which only AGL has anything
configured on, and even that is SIM/serial numbers, not IMEI) -
deliberately left out rather than guessed.
"""

import io
import time
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Always shown, in this order - the "who/what/when" a billing person
# actually needs. No asset/org id: those identify a MiX internal
# record, not something actionable for billing.
CORE_FIELDS = [
    {"key": "client", "label": "Client"},
    {"key": "fleetNumber", "label": "Fleet Number"},
    {"key": "plate", "label": "Plate / Registration"},
    {"key": "installedBy", "label": "Installed By"},
    {"key": "installDate", "label": "Install Date"},
    {"key": "installMonthDisplay", "label": "Install Month"},
]

# Tick/untick in the admin panel; included in the export only when
# explicitly selected (via the `fields` query param).
OPTIONAL_FIELDS = [
    {"key": "userState", "label": "State"},
    {"key": "make", "label": "Make"},
    {"key": "model", "label": "Model"},
    {"key": "year", "label": "Year"},
    {"key": "vin", "label": "VIN Number"},
    {"key": "odometer", "label": "Odometer"},
    {"key": "fuelType", "label": "Fuel Type"},
    {"key": "notes", "label": "Notes"},
]

_OPTIONAL_KEYS = {f["key"] for f in OPTIONAL_FIELDS}


def _parse_created_date(value):
    """MiX's CreatedDate is always an ISO 8601 UTC string
    ("2025-11-06T11:46:24Z") in every sample checked - not the old
    ASP.NET "/Date(...)/ " shape some other MiX fields use elsewhere in
    this app. Returns a naive datetime, or None if unparseable/missing."""
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%dT%H:%M:%SZ")
    except ValueError:
        return None


def month_display(month_str):
    """"2025-11" -> "November 2025". Falls back to the raw string if it
    doesn't parse, so a bad value never crashes the export."""
    if not month_str:
        return None
    try:
        return datetime.strptime(month_str, "%Y-%m").strftime("%B %Y")
    except ValueError:
        return month_str


def fetch_all_assets(client, org_ids, owner_by_org, inter_org_delay_seconds):
    """
    client: a MixApiClient. org_ids: the MiX orgs to poll (already
    resolved to whichever clients the admin selected).
    owner_by_org: {org_id: client_name}, from client_registry.platform_index.
    Returns (rows, errors) - errors keyed by org id, one org failing
    doesn't stop the others (same pattern as MixApiClient.get_assets_and_positions).
    """
    rows = []
    errors = {}
    for i, org_id in enumerate(org_ids):
        if i > 0:
            time.sleep(inter_org_delay_seconds)
        try:
            assets = client.get_assets(org_id)
        except Exception as e:
            logger.error(f"Asset install report: fetch failed for org {org_id}: {e}")
            errors[str(org_id)] = str(e)
            continue
        for asset in assets:
            installed = _parse_created_date(asset.get("CreatedDate"))
            rows.append({
                "client": owner_by_org.get(str(org_id), "Unassigned"),
                "fleetNumber": asset.get("FleetNumber") or "",
                "plate": asset.get("RegistrationNumber") or asset.get("Description") or "",
                "installedBy": asset.get("CreatedBy") or "",
                "installDate": installed.strftime("%d %b %Y") if installed else "Unknown",
                "installMonth": installed.strftime("%Y-%m") if installed else None,
                "installMonthDisplay": month_display(installed.strftime("%Y-%m")) if installed else "Unknown",
                "userState": asset.get("UserState") or "",
                "make": asset.get("Make") or "",
                "model": asset.get("Model") or "",
                "year": asset.get("Year") or "",
                "vin": asset.get("VinNumber") or "",
                "odometer": asset.get("Odometer"),
                "fuelType": asset.get("FuelType") or "",
                "notes": asset.get("Notes") or "",
            })
    return rows, errors


def build_workbook(rows, month=None, fields=None):
    """
    rows: the list from fetch_all_assets (or a cached snapshot's "rows").
    month: optional "YYYY-MM" filter - only assets whose parsed install
    date falls in that month are included, for the "what do I bill this
    month" use case.
    fields: optional list of OPTIONAL_FIELDS keys to include as extra
    columns, in OPTIONAL_FIELDS order - mirrors whichever checkboxes
    the admin ticked on screen.
    """
    if month:
        rows = [r for r in rows if r.get("installMonth") == month]
    rows = sorted(rows, key=lambda r: (r.get("client") or "", r.get("installMonth") or "", r.get("plate") or ""))

    selected_optional = [f for f in OPTIONAL_FIELDS if f["key"] in (fields or [])]
    columns = CORE_FIELDS + selected_optional

    wb = openpyxl.Workbook()
    ws = wb.active
    label = month_display(month)
    ws.title = (f"New Assets - {label}" if label else "All Assets")[:31]  # Excel's own 31-char sheet-name limit

    ws.append([c["label"] for c in columns])
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r.get(c["key"]) for c in columns])

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for i, c in enumerate(columns, start=1):
        col = get_column_letter(i)
        cell_lens = [len(str(ws.cell(row=r, column=i).value or "")) for r in range(2, ws.max_row + 1)]
        ws.column_dimensions[col].width = min(max([len(c["label"])] + cell_lens) + 2, 40)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
