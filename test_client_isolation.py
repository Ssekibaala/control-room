"""
Proves per-client data isolation at the only place that matters: the
actual bytes of the HTTP response.

The check is deliberately not "does the row list look right" but "does
another client's plate appear ANYWHERE in the serialized payload".
That distinction caught a real leak while this was being written: the
asset rows were correctly scoped, but the tampering sections
(tamperConfirmed/tamperUnconfirmed/qualityLog/tamperChecked) identify
vehicles by plate and carry no client field, so they were still handing
an unassigned technician 73 GTL plates. A row-shaped assertion would
have passed.

Run: python test_client_isolation.py
Needs data/fleet_today.json to contain rows for at least two clients.
"""

import os
import json

# Must be set before app is imported - importing it starts the live API
# pollers otherwise, which this test has no need for.
for _var in ("DISABLE_MIX_API_POLLER", "DISABLE_TELETRAC_API_POLLER", "DISABLE_FT_CLOUD_API_POLLER"):
    os.environ.setdefault(_var, "true")

import app as A  # noqa: E402
import permissions  # noqa: E402

DATA_PATH = os.path.join(os.path.dirname(__file__), "data", "fleet_today.json")


def _plates_by_client():
    with open(DATA_PATH) as f:
        data = json.load(f)
    out = {}
    for row in data.get("full", []):
        client = str(row.get("client", "")).strip()
        if client and row.get("plate"):
            out.setdefault(client, set()).add(row["plate"])
    return out


def _payload_for(role, assigned_clients):
    """One real request through the real route, with the Sheets-backed
    client lookup stubbed so the test needs no live accounts."""
    original = A._assigned_clients
    A._assigned_clients = lambda _username: assigned_clients
    A._user_clients_cache.clear()
    try:
        with A.app.test_client() as client:
            with client.session_transaction() as sess:
                sess["username"] = "isolation-test"
                sess["role"] = role
            return client.get("/api/dashboard-data").get_json()
    finally:
        A._assigned_clients = original
        A._user_clients_cache.clear()


def _plates_present(payload, plates):
    """How many of `plates` appear anywhere in the serialized response."""
    raw = json.dumps(payload)
    return sum(1 for p in plates if f'"{p}"' in raw)


def run():
    A.app.config["TESTING"] = True
    by_client = _plates_by_client()
    if len(by_client) < 2:
        print(f"SKIPPED - need two clients' data in {DATA_PATH}, found: {sorted(by_client)}")
        return True

    names = sorted(by_client)
    first, second = names[0], names[1]
    failures = []

    print(f"=== Client isolation ({len(names)} clients with data: {', '.join(names)}) ===")

    # `all` for admin rather than a fixed pair - the number of clients
    # grows as they're registered, and a test that assumes exactly two
    # starts failing on the third for no real reason.
    cases = [
        ("admin", [], set(names), "admin sees every client"),
        ("technician", [first], {first}, f"technician assigned {first}"),
        ("client", [second], {second}, f"client assigned {second}"),
        ("technician", [], set(), "technician with no assignment sees nothing"),
    ]

    for role, assigned, should_see, label in cases:
        payload = _payload_for(role, assigned)
        for name in names:
            found = _plates_present(payload, by_client[name])
            expected_any = name in should_see
            ok = (found > 0) if expected_any else (found == 0)
            verdict = "PASS" if ok else "FAIL"
            print(f"  [{verdict}] {label}: {name} plates in response = {found} "
                  f"({'expected' if expected_any else 'expected none'})")
            if not ok:
                failures.append(f"{label}: {name} plates present={found}, should_see={expected_any}")

        # meta.clients is registry-sourced, so it can legitimately list
        # clients that have no vehicles yet (just added, or their poll
        # hasn't run). The invariant that matters is directional: it must
        # cover everything they can see, and must never name a client
        # they can't.
        advertised = set(payload.get("meta", {}).get("clients") or [])
        missing = should_see - advertised
        forbidden = advertised & (set(names) - should_see)
        if missing or forbidden:
            print(f"  [FAIL] {label}: meta.clients={sorted(advertised)}; "
                  f"missing={sorted(missing)} forbidden={sorted(forbidden)}")
            failures.append(f"{label}: meta.clients missing={sorted(missing)} forbidden={sorted(forbidden)}")
        else:
            print(f"  [PASS] {label}: meta.clients={sorted(advertised)}")

    print()
    print("=== Every endpoint that takes a plate or serves bulk data ===")
    # These are the paths that bypass /api/dashboard-data entirely. Each
    # one was verified leaking before it was fixed, so each is checked
    # here rather than assumed to follow from the row filtering above.
    mine, theirs = first, second
    my_plate = sorted(by_client[mine])[0]
    their_plate = sorted(by_client[theirs])[0]
    other_plates = set().union(*(by_client[n] for n in names if n != mine))

    original = A._assigned_clients
    A._assigned_clients = lambda _u: [mine]
    A._user_clients_cache.clear()
    try:
        with A.app.test_client() as client:
            with client.session_transaction() as sess:
                sess["username"] = "isolation-test"
                sess["role"] = "technician"

            checks = [
                ("feedback-history, another client's vehicle",
                 lambda: client.get(f"/api/feedback-history/{their_plate}").status_code, 404),
                ("feedback-history, own vehicle still readable",
                 lambda: client.get(f"/api/feedback-history/{my_plate}").status_code, 200),
                ("feedback POST onto another client's vehicle",
                 lambda: client.post("/api/feedback", json={
                     "plate": their_plate, "comment": "isolation probe",
                     "reportedBy": "isolation-test", "requiresFollowup": False}).status_code, 404),
                ("tamper-check on another client's vehicle",
                 lambda: client.post("/api/tamper-check", json={
                     "plate": their_plate, "comment": "isolation probe"}).status_code, 404),
            ]
            for label, run, expected in checks:
                got = run()
                ok = got == expected
                print(f"  [{'PASS' if ok else 'FAIL'}] {label}: {got} (expected {expected})")
                if not ok:
                    failures.append(f"{label}: got {got}, expected {expected}")

            # Bulk payloads: assert on the serialized bytes, since these
            # embed plates in shapes the row filters never touch.
            for path in ("/api/mix/snapshot", "/api/teletrac/snapshot", "/api/ftcloud/snapshot"):
                body = client.get(path).data.decode("utf-8", "ignore")
                hits = sum(1 for p in other_plates if f'"{p}"' in body)
                print(f"  [{'PASS' if not hits else 'FAIL'}] {path}: other-client plates = {hits}")
                if hits:
                    failures.append(f"{path} leaked {hits} plates")

            listed = [c["name"] for c in (client.get("/api/clients").get_json() or {}).get("clients", [])]
            ok = listed == [mine]
            print(f"  [{'PASS' if ok else 'FAIL'}] /api/clients: {listed} (expected only {mine})")
            if not ok:
                failures.append(f"/api/clients returned {listed}")

            # The export must be BOTH scoped and still complete - a
            # filter that empties the workbook would pass a leak test
            # while breaking the feature.
            resp = client.get("/api/export/integrity")
            if resp.status_code != 200:
                print(f"  [FAIL] export/integrity: status {resp.status_code}")
                failures.append(f"export returned {resp.status_code}")
            else:
                import io
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(resp.data))
                seen = set()
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                seen.add(cell.strip())
                leaked = len(seen & other_plates)
                kept = len(seen & by_client[mine])
                ok = leaked == 0 and kept > 0
                print(f"  [{'PASS' if ok else 'FAIL'}] export/integrity: {kept} own plates kept, "
                      f"{leaked} other-client plates leaked")
                if not ok:
                    failures.append(f"export kept={kept} leaked={leaked}")
    finally:
        A._assigned_clients = original
        A._user_clients_cache.clear()

    print()
    print("=== Notification routing is per client ===")
    import notifications
    for name in names:
        recips = set(notifications._client_recipients(name))
        others = set()
        for other in names:
            if other != name:
                others |= set(notifications._client_recipients(other))
        # A shared mailbox legitimately serving two clients is possible,
        # so this only asserts the routing is not the old "everyone"
        # behaviour: recipients must be a per-client lookup, not the
        # union of every client's contacts.
        everyone = set(notifications._client_recipients(None))
        ok = recips <= everyone and (not everyone or recips != everyone or len(names) == 1)
        print(f"  [{'PASS' if ok else 'FAIL'}] {name}: {sorted(recips) or 'no contacts on file'}")
        if not ok:
            failures.append(f"{name} recipients equal the all-client union")

    print()
    print("=== Nobody can grant access they don't hold ===")
    grant_cases = [
        ("technician", [first], [second], False, f"technician with {first} granting {second}"),
        ("technician", [first], [first], True, f"technician with {first} granting {first}"),
        ("admin", [], [first, second], True, "admin granting anything"),
    ]
    for role, held, requested, should_allow, label in grant_cases:
        ok, disallowed = permissions.can_grant_clients(role, held, requested)
        verdict = "PASS" if ok is should_allow else "FAIL"
        print(f"  [{verdict}] {label}: allowed={ok} refused={disallowed}")
        if ok is not should_allow:
            failures.append(label)

    print()
    print("=== An empty assignment means nothing, never everything ===")
    if permissions.visible_clients("client", []) != set():
        failures.append("empty assignment did not resolve to an empty client set")
        print("  [FAIL] empty assignment should resolve to no clients")
    else:
        print("  [PASS] empty assignment resolves to no clients")

    print()
    if failures:
        print(f"RESULT: FAILED -> {failures}")
        return False
    print("RESULT: PASSED - client isolation holds in the actual HTTP response bytes.")
    return True


if __name__ == "__main__":
    import sys
    sys.exit(0 if run() else 1)
